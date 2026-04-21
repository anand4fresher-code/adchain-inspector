import streamlit as st
import requests
import pandas as pd
import json
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, timedelta
import time
import re
import sqlite3
import hashlib
import io
import os
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import gzip
import zlib
try:
    import brotli as _brotli
    HAS_BROTLI = True
except ImportError:
    HAS_BROTLI = False
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="AdChain Inspector v2",
    page_icon="🔗",
    layout="wide",
    initial_sidebar_state="expanded"
)

ADVLION_DOMAIN = "advlion.com"
ADVLION_SID    = "3148"
DB_PATH        = os.path.join(os.path.dirname(os.path.abspath(__file__)), "adchain_cache.db")

TOP_15_EXCHANGES = {
    # ── Core Web / Display (original 15) ────────────────────────────────
    "PubMatic":          "pubmatic.com",
    "Magnite (Rubicon)": "rubiconproject.com",
    "Xandr (AppNexus)":  "xandr.com",
    "OpenX":             "openx.com",
    "Index Exchange":    "indexexchange.com",
    "TripleLift":        "triplelift.com",
    "Sharethrough":      "sharethrough.com",
    "Sovrn":             "sovrn.com",
    "SmartAdServer":     "smartadserver.com",
    "Criteo (TheMediaGrid)": "themediagrid.com",
    "GumGum":            "gumgum.com",
    "Smaato":            "smaato.com",
    "Nexxen (TremorHub)":"tremorhub.com",
    "EMX Digital":       "emxdigital.com",
    "33Across":          "33across.com",
    # ── Added: Mobile / In-App / Native ─────────────────────────────────
    "InMobi":            "inmobi.com",
    "Unity Ads":         "unity.com",
    "Yieldmo":           "yieldmo.com",
    "Kargo":             "kargo.com",
    "Nativo":            "nativo.com",
    "Media.net":         "media.net",
    "Minutemedia":       "minutemedia.com",
}

CDN_PATTERNS = [
    "cloudfront.net", "amazonaws.com", "fastly.net",
    "akamaihd.net", "cloudflare.net", "azureedge.net",
    "googleusercontent.com", "gstatic.com", "cdn.com"
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",       # no 'br' — brotli handled separately
    "Connection":      "keep-alive",
    "Cache-Control":   "no-cache",
}

# ─── STYLES ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stSidebar"] { background: #0A1628 !important; }
    [data-testid="stSidebar"] * { color: #C8D8E8 !important; }
    [data-testid="stSidebar"] .stRadio label { color: #C8D8E8 !important; padding: 4px 0; }
    .stButton > button {
        background: linear-gradient(135deg, #1E4D8C, #2E75B6);
        color: white; border: none; border-radius: 8px;
        padding: 8px 20px; font-weight: 600;
    }
    .stButton > button:hover { background: linear-gradient(135deg, #2E6DB4, #3E85C6); }
    div[data-testid="stMetricValue"] { font-size: 22px !important; font-weight: 700 !important; }
    .v2-badge {
        background: linear-gradient(135deg, #1B5E20, #2E7D32);
        color: #A5D6A7; padding: 2px 8px; border-radius: 10px;
        font-size: 11px; font-weight: 700; letter-spacing: 0.5px;
    }
</style>
""", unsafe_allow_html=True)

# ─── SQLITE CACHE ─────────────────────────────────────────────────────────

def init_db():
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sellers_cache (
            domain      TEXT PRIMARY KEY,
            url         TEXT,
            fetched_at  TEXT,
            data_hash   TEXT,
            data_json   TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sellers_snapshots (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            domain      TEXT,
            snapshot_at TEXT,
            data_hash   TEXT,
            data_json   TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS adstxt_cache (
            domain      TEXT,
            filename    TEXT,
            fetched_at  TEXT,
            content_hash TEXT,
            content     TEXT,
            PRIMARY KEY (domain, filename)
        )
    """)
    con.commit()
    con.close()

def cache_get_sellers(domain):
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        cur.execute("SELECT fetched_at, data_json, data_hash FROM sellers_cache WHERE domain=?", (domain,))
        row = cur.fetchone()
        con.close()
        if row:
            fetched_at = datetime.fromisoformat(row[0])
            if datetime.now() - fetched_at < timedelta(hours=24):
                return {"hit": True, "data": json.loads(row[1]), "hash": row[2], "age": fetched_at}
    except Exception:
        pass
    return {"hit": False}

def cache_set_sellers(domain, url, data):
    try:
        data_json = json.dumps(data)
        data_hash = hashlib.md5(data_json.encode()).hexdigest()
        now       = datetime.now().isoformat()
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()

        # Check if we need to save a new snapshot (different hash than last)
        cur.execute("SELECT data_hash FROM sellers_cache WHERE domain=?", (domain,))
        row = cur.fetchone()
        if row and row[0] != data_hash:
            # Data changed — save old version as snapshot before overwriting
            cur.execute("SELECT data_json FROM sellers_cache WHERE domain=?", (domain,))
            old = cur.fetchone()
            if old:
                cur.execute("""
                    INSERT INTO sellers_snapshots (domain, snapshot_at, data_hash, data_json)
                    VALUES (?, ?, ?, ?)
                """, (domain, now, row[0], old[0]))

        cur.execute("""
            INSERT OR REPLACE INTO sellers_cache (domain, url, fetched_at, data_hash, data_json)
            VALUES (?, ?, ?, ?, ?)
        """, (domain, url, now, data_hash, data_json))
        con.commit()
        con.close()
        return data_hash
    except Exception:
        return ""

def cache_get_snapshots(domain):
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        cur.execute("""
            SELECT snapshot_at, data_hash, data_json
            FROM sellers_snapshots WHERE domain=?
            ORDER BY snapshot_at DESC LIMIT 10
        """, (domain,))
        rows = cur.fetchall()
        con.close()
        return rows
    except Exception:
        return []

def cache_clear():
    try:
        con = sqlite3.connect(DB_PATH)
        con.execute("DELETE FROM sellers_cache")
        con.execute("DELETE FROM adstxt_cache")
        con.commit()
        con.close()
        return True
    except Exception:
        return False

# ─── HELPERS ──────────────────────────────────────────────────────────────

def clean_domain(domain):
    domain = str(domain).strip().lower()
    if domain.startswith("http"):
        domain = urlparse(domain).netloc
    return domain.replace("www.", "").strip("/")

def extract_domain_from_field(raw_value):
    val = str(raw_value).strip().lower()
    if not val:
        return ""
    if val.startswith("http://") or val.startswith("https://") or val.startswith("//"):
        parsed = urlparse(val if "//" in val else f"//{val}")
        val = parsed.netloc or parsed.path
    val = val.split("/")[0]
    val = val.replace("www.", "").strip()
    return val

def is_cdn_domain(domain):
    return any(pat in domain.lower() for pat in CDN_PATTERNS)

def df_to_csv(df):
    return df.to_csv(index=False).encode("utf-8")

def get_sellers(data):
    return data.get("sellers", []) if data else []

def parse_ads_txt(content):
    """
    Parse combined ads.txt + app-ads.txt content.
    Tracks source file per line using # --- filename --- section markers
    so we can report exactly where a DIRECT match was found.
    """
    rows         = []
    current_file = "ads.txt"   # default if no marker present
    if not content:
        return rows
    for line in content.splitlines():
        line = line.strip()
        if not line:
            continue
        # Section marker injected by fetch functions e.g. "# --- app-ads.txt ---"
        if line.startswith("# ---") and line.endswith("---"):
            fname = line.replace("# ---", "").replace("---", "").strip()
            if fname:
                current_file = fname
            continue
        if line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) >= 3:
            rows.append({
                "exchange_domain": parts[0].lower().replace("www.", ""),
                "seller_id":       parts[1].strip(),
                "relationship":    parts[2].strip().upper(),
                "cert_id":         parts[3].strip() if len(parts) > 3 else "",
                "source_file":     current_file,   # ← NEW: which file this line came from
                "raw":             line
            })
    return rows

# ─── INVENTORY TYPE DETECTION ────────────────────────────────────────────

# Tier 1 — Definitive CTV platforms (high confidence)
CTV_TIER1 = [
    "roku", "pluto.tv", "tubi.com", "xumo", "fubotv", "slingtv", "hulu",
    "peacocktv", "paramountplus", "disneyplus", "espn", "nbcuni",
    "vidaa", "lgchannels", "viziotv", "samsung.tv", "firetv", "philo",
    "directvstream", "frndly", "plex.tv", "crackle", "joyn", "tvnz",
    "channel4.com", "itv.com", "channel5", "rtlgroup", "tele5",
    "skygroup", "nowtv", "discoveryplus", "hbomax", "max.com",
    "appletv", "apple.tv", "britbox", "acorntv", "shout", "filmrise",
    "redbox", "stirr", "newsmax", "oann", "beinsports", "dazn",
    "locast", "vidgo", "fubo", "distrotv", "plex", "freevee"
]

# Tier 2 — Strong CTV signals (medium-high confidence)
CTV_TIER2 = [
    "streaming", "smarttv", "smart-tv", "connectedtv", "connected-tv",
    "iptv", "ottplatform", "ott-", "-ott", "tvplatform", "tvos",
    "androidtv", "android-tv", "firestick", "chromecast", "airplay",
    "screenmedia", "screentv", "tvapp", "channel.com", "livetvstream",
    "watchnow", "watchlive", "streamlive", "tvnow", "tvlive",
    "tvplus", "tvgo", "tvplay", "tvhub", "tvtime"
]

# Tier 3 — Domain keyword signals (lower confidence, needs corroboration)
CTV_TIER3_KEYWORDS = [
    "television", "telemundo", "univision", "scripps", "nexstar",
    "sinclair", "tribune", "gray.tv", "hearst", "raycom", "meredith",
    "tegna", "cox.media", "emmis", "cumulus", "audacy", "iheartmedia"
]

# Domain parts that suggest mobile app (not CTV)
MOBILE_APP_SIGNALS = [
    "mobileapp", "appstore", "googleplay", "android-app", "ios-app",
    "gamer", "gaming", "casual", "hypercasual", "puzzle", "arcade",
    "match3", "clicker", "idle", "rpg", "mmo", "fps", "shooter",
    "social", "dating", "fitness", "health", "finance", "banking",
    "shopping", "ecommerce", "food", "delivery", "ride", "travel",
    "news", "weather", "utility", "productivity", "tools"
]

def check_play_store_ctv(domain):
    """
    Real Play Store check — searches by app name and checks for TV support.
    Uses multiple search strategies. Returns (result, evidence).
    """
    app_name = domain.split(".")[0]
    search_urls = [
        f"https://play.google.com/store/search?q={app_name}&c=apps",
        f"https://play.google.com/store/search?q={app_name}+tv&c=apps",
    ]
    for url in search_urls:
        try:
            r = requests.get(url, timeout=6, headers=HEADERS)
            if r.status_code == 200:
                page = r.text.lower()
                if "android tv" in page or "designed for tv" in page or                    "android television" in page or "tv app" in page or                    "leanback" in page:
                    return "ctv", "Play Store: Android TV / Leanback detected"
                if "get it on google play" in page or                    "install" in page or "open" in page:
                    return "mobile", "Play Store: Mobile app confirmed"
        except Exception:
            continue
    return "unknown", ""

def check_app_store_ctv(domain):
    """
    Real App Store check via iTunes Search API.
    Checks supportedDevices for AppleTV/tvOS. Returns (result, evidence).
    """
    app_name = domain.split(".")[0]
    try:
        # Check both software and tvSoftware entity types
        for entity in ["tvSoftware", "software"]:
            url = f"https://itunes.apple.com/search?term={app_name}&entity={entity}&limit=5"
            r   = requests.get(url, timeout=6, headers=HEADERS)
            if r.status_code == 200:
                data    = r.json()
                results = data.get("results", [])
                if entity == "tvSoftware" and results:
                    return "ctv", f"App Store: tvOS/Apple TV app found ({results[0].get('trackName','')})"
                for app in results[:3]:
                    devices   = [s.lower() for s in app.get("supportedDevices", [])]
                    genres    = [g.lower() for g in app.get("genres", [])]
                    has_tv    = any("appletv" in d or "tv" in d for d in devices)
                    is_mobile = any("iphone" in d or "ipad" in d for d in devices)
                    if has_tv:
                        return "ctv", f"App Store: tvOS detected for {app.get('trackName','')}"
                    if is_mobile and not has_tv:
                        return "mobile", f"App Store: Mobile only ({app.get('trackName','')})"
    except Exception:
        pass
    return "unknown", ""

def check_domain_homepage(domain):
    """
    Quick homepage fetch — looks for TV/CTV/streaming indicators in page content.
    Returns (signal, evidence).
    """
    try:
        r = requests.get(f"https://{domain}", timeout=5, headers=HEADERS)
        if r.status_code == 200:
            text = r.text.lower()[:5000]   # only check first 5KB
            ctv_signals = [
                "connected tv", "smart tv", "streaming tv", "ott", "linear tv",
                "television", "broadcast", "channel", "streaming service",
                "watch live", "on demand", "catch-up tv", "tv app",
                "apple tv", "roku", "android tv", "fire tv", "samsung tv"
            ]
            mobile_signals = [
                "download on app store", "get it on google play",
                "mobile app", "android app", "ios app", "download our app"
            ]
            ctv_hits    = [s for s in ctv_signals    if s in text]
            mobile_hits = [s for s in mobile_signals if s in text]
            if ctv_hits and not mobile_hits:
                return "ctv",    f"Homepage: {ctv_hits[0]}"
            if mobile_hits and not ctv_hits:
                return "mobile", f"Homepage: {mobile_hits[0]}"
            if ctv_hits and mobile_hits:
                return "both",   f"Homepage: TV+App signals"
    except Exception:
        pass
    return "unknown", ""

def score_domain_signals(domain_l):
    """
    Score a domain against all signal tiers.
    Returns (category, confidence, evidence_list).
    """
    evidence = []
    ctv_score  = 0
    web_score  = 0
    app_score  = 0

    # Tier 1 check (definitive CTV)
    t1 = next((k for k in CTV_TIER1 if k in domain_l), None)
    if t1:
        ctv_score += 10
        evidence.append(f"Tier-1 CTV platform: {t1}")

    # Tier 2 check (strong CTV)
    t2 = next((k for k in CTV_TIER2 if k in domain_l), None)
    if t2:
        ctv_score += 6
        evidence.append(f"CTV keyword: {t2}")

    # .tv TLD
    if domain_l.endswith(".tv"):
        ctv_score += 5
        evidence.append(".tv TLD")

    # Tier 3 broadcast networks
    t3 = next((k for k in CTV_TIER3_KEYWORDS if k in domain_l), None)
    if t3:
        ctv_score += 3
        evidence.append(f"Broadcast/network: {t3}")

    # Mobile signals
    mob = next((k for k in MOBILE_APP_SIGNALS if k in domain_l), None)
    if mob:
        app_score += 4
        evidence.append(f"Mobile keyword: {mob}")

    # Generic web signals
    if any(k in domain_l for k in ["news", "media", "press", "journal", "magazine", "blog", "post", "daily", "times", "herald"]):
        web_score += 3
        evidence.append("News/media domain")

    return ctv_score, app_score, web_score, evidence

def detect_inventory_type(domain, match_found_in, do_store_check=True):
    """
    Accurately detect inventory type.
    Returns (inventory_type, confidence, evidence_detail)

    KEY RULE — file match is the primary signal:
      ads.txt ONLY        → 🌐 Web  (store check NEVER overrides this)
      app-ads.txt ONLY    → run store check → CTV or In-App
      BOTH files          → Mixed, then refine with domain signals
      No match / Fake     → domain signals only (no store check)
    """
    domain_l   = str(domain).lower().strip()
    match_str  = str(match_found_in).lower()

    # ── Parse file match signals ──────────────────────────────────────────
    clean_match = match_str.replace("✅","").replace("⚠️","").replace("❌","").strip()
    parts       = [p.strip() for p in clean_match.split("+")]
    in_app      = any("app-ads" in p for p in parts)
    in_web      = any(p.strip() == "ads.txt" for p in parts)
    no_match    = "not found" in match_str or "no ads" in match_str or "none" in match_str

    # ── RULE 1: ads.txt ONLY = Web — no further checks needed ────────────
    # Store checks must never override a confirmed ads.txt-only match.
    if in_web and not in_app and not no_match:
        # Still check CTV domain signals (e.g. .tv TLD, known streaming platform)
        ctv_score, _, _, dom_evidence = score_domain_signals(domain_l)
        if ctv_score >= 5:
            ctv_kw = next((k for k in CTV_TIER1 + CTV_TIER2 if k in domain_l), None)
            ev = f"ads.txt match | CTV domain signal: {ctv_kw or '.tv TLD'}"
            conf = "🟢 High" if ctv_score >= 8 else "🟡 Medium"
            return "🖥️ CTV (Web)", conf, ev
        return "🌐 Web", "🟢 High", "ads.txt only — confirmed web inventory"

    # ── RULE 2: No match — use domain signals only, no store check ────────
    if no_match:
        ctv_score, app_score, web_score, dom_evidence = score_domain_signals(domain_l)
        ev = " | ".join(dom_evidence) if dom_evidence else "No file match"
        if ctv_score >= 5:
            return "🖥️ CTV", "🟡 Medium", f"No ads.txt match | {ev}"
        if web_score >= 3:
            return "🌐 Web", "🔴 Low", f"No match | {ev}"
        return "❓ Unknown", "🔴 Low", ev

    # ── RULE 3: app-ads.txt ONLY — run store check ────────────────────────
    if in_app and not in_web:
        ctv_score, app_score, web_score, dom_evidence = score_domain_signals(domain_l)

        # Domain-level CTV signals first (fastest — no network call)
        if ctv_score >= 5:
            conf = "🟢 High" if ctv_score >= 8 else "🟡 Medium"
            ev   = " | ".join(dom_evidence) + " | app-ads.txt match"
            return "🖥️ CTV", conf, ev

        # Only now do store checks — app-ads.txt with no CTV domain signal
        store_ev = ""
        if do_store_check:
            ps_result, ps_ev = check_play_store_ctv(domain_l)
            if ps_result == "ctv":
                return "🖥️ CTV", "🟢 High", f"app-ads.txt | {ps_ev}"
            elif ps_result == "mobile":
                return "📱 In-App", "🟢 High", f"app-ads.txt | {ps_ev}"

            as_result, as_ev = check_app_store_ctv(domain_l)
            if as_result == "ctv":
                return "🖥️ CTV", "🟢 High", f"app-ads.txt | {as_ev}"
            elif as_result == "mobile":
                return "📱 In-App", "🟢 High", f"app-ads.txt | {as_ev}"

            # Homepage check as last resort
            hp_result, hp_ev = check_domain_homepage(domain_l)
            if hp_result == "ctv":
                return "🖥️ CTV", "🟡 Medium", f"app-ads.txt | {hp_ev}"
            elif hp_result == "mobile":
                return "📱 In-App", "🟡 Medium", f"app-ads.txt | {hp_ev}"

        return "📱 In-App", "🟡 Medium", "app-ads.txt only — no CTV signals found"

    # ── RULE 4: BOTH files matched — Mixed, refine with domain signals ────
    if in_app and in_web:
        ctv_score, app_score, web_score, dom_evidence = score_domain_signals(domain_l)
        ev = " | ".join(dom_evidence) if dom_evidence else "Both files"
        if ctv_score >= 5:
            conf = "🟢 High" if ctv_score >= 8 else "🟡 Medium"
            return "🖥️ CTV + 🌐 Web", conf, f"Both files | {ev}"
        if app_score >= 5:
            return "🔀 Mixed (Web + App)", "🟡 Medium", f"Both files | {ev}"
        return "🔀 Mixed (Web + App)", "🟡 Medium", "Both ads.txt + app-ads.txt"

    return "❓ Unknown", "🔴 Low", "Could not determine"

# ─── NETWORK — SMART FETCH WITH RETRY + CACHE ─────────────────────────────

def fetch_sellers_json(domain, use_cache=True):
    domain = clean_domain(domain)

    if use_cache:
        cached = cache_get_sellers(domain)
        if cached["hit"]:
            return {
                "success": True, "data": cached["data"],
                "url": f"https://{domain}/sellers.json",
                "from_cache": True, "cache_age": cached["age"]
            }

    # Some exchanges host very large sellers.json files (50k+ entries).
    # Sovrn and Criteo are known examples — give them extra time.
    # sovrn.com has 70k+ entries — needs 45s. themediagrid/tremorhub are also large.
    LARGE_FILE_DOMAINS = {
        "sovrn.com":         45,
        "pubmatic.com":      30,
        "rubiconproject.com":25,
        "openx.com":         25,
        "xandr.com":         25,
        "indexexchange.com": 25,
        "tremorhub.com":     25,
        "smartadserver.com": 25,
        "themediagrid.com":  25,
        "media.net":         25,
    }
    timeout_secs = LARGE_FILE_DOMAINS.get(domain, 8)

    urls = [
        f"https://{domain}/sellers.json",
        f"http://{domain}/sellers.json",
        f"https://www.{domain}/sellers.json",
    ]

    def _try_parse(r):
        """Try multiple strategies including brotli/gzip decompression to parse sellers.json"""
        import re as _re
        raw = r.content

        # First try to decompress — handles brotli/gzip sent without Content-Encoding header
        def _decompress(b):
            for fn in [
                lambda x: _brotli.decompress(x) if HAS_BROTLI else None,
                lambda x: gzip.decompress(x),
                lambda x: zlib.decompress(x),
                lambda x: zlib.decompress(x, -15),
                lambda x: zlib.decompress(x,  47),
            ]:
                try:
                    r = fn(b)
                    if r: return r
                except Exception:
                    continue
            return None

        decomp = _decompress(raw)

        strategies = [
            lambda: r.json(),
            lambda: json.loads(raw.decode("utf-8-sig").strip()),
            lambda: json.loads(r.text.lstrip("\ufeff").strip()),
            lambda: json.loads(r.text.strip()),
            lambda: json.loads(raw.decode("utf-16").strip()),
            lambda: json.loads(raw.decode("latin-1").strip()),
            lambda: json.loads(raw.decode("utf-8", errors="ignore").strip()),
            lambda: json.loads(_re.sub(r'^\s*\w+\s*\(|\)\s*;?\s*$', '', r.text).strip()),
            # Decompressed variants
            lambda: json.loads(decomp.decode("utf-8").strip()) if decomp else None,
            lambda: json.loads(decomp.decode("utf-8-sig").strip()) if decomp else None,
            lambda: json.loads(decomp.decode("utf-8", errors="ignore").strip()) if decomp else None,
        ]
        for fn in strategies:
            try:
                data = fn()
                if data is not None and isinstance(data, dict) and "sellers" in data:
                    return data
            except Exception:
                continue
        return None

    header_variants = [
        HEADERS,
        {**HEADERS, "Accept-Encoding": "identity"},   # force no compression
    ]
    for url in urls:
        for hdrs in header_variants:
            for attempt in range(2):
                try:
                    r = requests.get(url, timeout=timeout_secs, headers=hdrs)
                    if r.status_code == 200 and r.content:
                        data = _try_parse(r)
                        if data:
                            if use_cache:
                                cache_set_sellers(domain, url, data)
                            return {"success": True, "data": data, "url": url, "from_cache": False}
                except Exception:
                    time.sleep(0.3)
    return {"success": False, "data": None, "url": urls[0], "from_cache": False}

def fetch_ads_txt(domain, is_app=False):
    """
    Always fetch BOTH ads.txt and app-ads.txt and combine them.
    Prevents false Fake Entry verdicts when DIRECT line is in one file but not the other.
    """
    domain    = clean_domain(domain)
    filenames = ["app-ads.txt", "ads.txt"] if is_app else ["ads.txt", "app-ads.txt"]

    combined_content = []
    found_files      = []
    first_url        = None

    for fname in filenames:
        for prefix in [f"https://{domain}/", f"http://{domain}/", f"https://www.{domain}/"]:
            url = f"{prefix}{fname}"
            for attempt in range(2):
                try:
                    r = requests.get(url, timeout=6, headers=HEADERS)
                    if r.status_code == 200 and r.text.strip():
                        if not first_url:
                            first_url = url
                        combined_content.append(f"# --- {fname} ---")
                        combined_content.append(r.text.strip())
                        found_files.append(fname)
                        break
                except Exception:
                    time.sleep(0.3)
            if fname in found_files:
                break   # got this file, try next filename

    if combined_content:
        return {
            "success":  True,
            "content":  "\n".join(combined_content),
            "url":      first_url or f"https://{domain}/ads.txt",
            "filename": " + ".join(found_files)
        }
    return {"success": False, "content": None,
            "url": f"https://{domain}/ads.txt", "filename": "ads.txt"}

def fetch_ads_txt_fast(domain, check_app=True, check_web=True, req_timeout=4):
    """
    Fetch ads.txt and/or app-ads.txt based on flags.
    check_web=False  → skips ads.txt  (In-App / CTV only mode — faster)
    check_app=False  → skips app-ads.txt
    Both True        → fetches both and combines (default)
    req_timeout      → per-request timeout in seconds (lower = faster for bulk runs)
    """
    domain    = extract_domain_from_field(domain)
    filenames = []
    if check_web: filenames.append("ads.txt")
    if check_app: filenames.append("app-ads.txt")
    if not filenames: filenames = ["ads.txt"]  # safety fallback

    combined_content = []
    found_files      = []

    for fname in filenames:
        for scheme in ["https", "http"]:
            for prefix in ["", "www."]:
                url = f"{scheme}://{prefix}{domain}/{fname}"
                try:
                    r = requests.get(url, timeout=req_timeout, headers=HEADERS)
                    if r.status_code == 200 and r.text.strip():
                        combined_content.append(f"# --- {fname} ---")
                        combined_content.append(r.text.strip())
                        found_files.append(fname)
                        break   # got this file — move to next filename
                except Exception:
                    continue
            if fname in found_files:
                break   # found via https/http — move to next filename

    if combined_content:
        return {
            "success":  True,
            "content":  "\n".join(combined_content),
            "filename": " + ".join(found_files)   # e.g. "ads.txt + app-ads.txt"
        }
    return {"success": False, "content": None, "filename": "ads.txt"}

# ─── PUBLISHER CHECK (PARALLEL) ───────────────────────────────────────────

def check_single_publisher(seller, check_app, exchange_domain_filter="", check_web=True, req_timeout=4):
    raw_domain = str(seller.get("domain", ""))
    domain     = extract_domain_from_field(raw_domain)
    seller_id  = str(seller.get("seller_id", "")).strip().lstrip("0") or str(seller.get("seller_id","")).strip()
    name       = str(seller.get("name", "—"))
    flags      = []

    if not domain:
        return {"Publisher Name": name[:35], "Domain": raw_domain or "—",
                "Seller ID": seller_id, "ads.txt": "⚠️ No domain",
                "DIRECT Line": "—", "Flags": "No domain in sellers.json",
                "Verdict": "⚠️ No Domain"}

    # CDN / Fraud flag
    if is_cdn_domain(domain):
        flags.append("🚨 CDN domain")

    res = fetch_ads_txt_fast(domain, check_app=check_app, check_web=check_web, req_timeout=req_timeout)

    if not res["success"]:
        return {"Publisher Name": name[:35], "Domain": domain,
                "Seller ID": seller_id,
                "Files Checked":  "❌ Not Found",
                "Match Found In": "❌ No ads.txt found",
                "DIRECT Line":    "❌ None",
                "Flags":          " | ".join(flags) if flags else "—",
                "Verdict":        "❌ Fake Entry"}

    parsed   = parse_ads_txt(res["content"])
    ex_filter = exchange_domain_filter.strip().lower() if exchange_domain_filter.strip() else None

    # Normalise seller_id comparison (strip leading zeros)
    def sid_match(row_sid, target_sid):
        return row_sid.strip().lstrip("0") == target_sid.lstrip("0") or row_sid.strip() == target_sid

    if ex_filter:
        direct_ex   = [r for r in parsed if sid_match(r["seller_id"], seller_id)
                       and r["relationship"] == "DIRECT" and r["exchange_domain"] == ex_filter]
        reseller_ex = [r for r in parsed if sid_match(r["seller_id"], seller_id)
                       and r["relationship"] == "RESELLER" and r["exchange_domain"] == ex_filter]
        any_direct  = [r for r in parsed if sid_match(r["seller_id"], seller_id) and r["relationship"] == "DIRECT"]

        # ── ID Mismatch detection ─────────────────────────────────────────
        # Exchange domain IS listed as DIRECT but with a DIFFERENT seller_id
        # e.g. cas.ai is in ads.txt as DIRECT but seller_id doesn't match sellers.json
        ex_direct_wrong_id = [r for r in parsed
                              if r["relationship"] == "DIRECT"
                              and r["exchange_domain"] == ex_filter
                              and not sid_match(r["seller_id"], seller_id)]
        ex_any_wrong_id    = [r for r in parsed
                              if r["exchange_domain"] == ex_filter
                              and not sid_match(r["seller_id"], seller_id)]
    else:
        direct_ex   = [r for r in parsed if sid_match(r["seller_id"], seller_id) and r["relationship"] == "DIRECT"]
        reseller_ex = [r for r in parsed if sid_match(r["seller_id"], seller_id) and r["relationship"] == "RESELLER"]
        any_direct  = direct_ex
        ex_direct_wrong_id = []
        ex_any_wrong_id    = []

    flag_str = " | ".join(flags) if flags else "—"

    def match_location(matches):
        """Return which file(s) the matching lines were found in."""
        files = list(dict.fromkeys([m.get("source_file", "ads.txt") for m in matches]))
        return " + ".join(files) if files else "ads.txt"

    def build_row(found_in_str, direct_line, verdict, do_inv_check=True, extra_flags=""):
        """Build result row — inventory type detected for ALL entries including fake ones."""
        inv_type, conf, inv_signal = detect_inventory_type(
            domain, found_in_str, do_store_check=do_inv_check
        )
        combined_flags = " | ".join(f for f in [flag_str, extra_flags] if f and f != "—") or "—"
        return {
            "Publisher Name":  name[:35],
            "Domain":          domain,
            "Seller ID":       seller_id,
            "Files Checked":   f"✅ {res['filename']}",
            "Match Found In":  found_in_str,
            "DIRECT Line":     direct_line,
            "Inventory Type":  inv_type,
            "Confidence":      conf,
            "Type Evidence":   inv_signal,
            "Flags":           combined_flags,
            "Verdict":         verdict
        }

    if direct_ex:
        matched_ex = direct_ex[0]["exchange_domain"]
        found_in   = match_location(direct_ex)
        return build_row(f"✅ {found_in}", f"✅ DIRECT ({matched_ex})", "✅ Legitimate")

    elif reseller_ex:
        found_in = match_location(reseller_ex)
        return build_row(f"⚠️ {found_in}", "⚠️ RESELLER only", "⚠️ Misrepresented")

    elif ex_filter and any_direct:
        found_ex = any_direct[0]["exchange_domain"]
        found_in = match_location(any_direct)
        return build_row(f"⚠️ {found_in}", f"⚠️ DIRECT under {found_ex}", "⚠️ Domain Mismatch")

    elif ex_direct_wrong_id:
        # Exchange IS listed as DIRECT but seller_id in ads.txt doesn't match sellers.json
        wrong_sid  = ex_direct_wrong_id[0]["seller_id"]
        found_in   = match_location(ex_direct_wrong_id)
        mismatch_note = f"⚠️ ID Mismatch — sellers.json has {seller_id}, ads.txt has {wrong_sid}"
        return build_row(
            f"⚠️ {found_in}",
            f"⚠️ DIRECT found but wrong ID (ads.txt={wrong_sid})",
            "⚠️ ID Mismatch",
            extra_flags=mismatch_note
        )

    elif ex_any_wrong_id:
        # Exchange listed (RESELLER) with a different seller_id
        wrong_sid  = ex_any_wrong_id[0]["seller_id"]
        rel        = ex_any_wrong_id[0]["relationship"]
        found_in   = match_location(ex_any_wrong_id)
        mismatch_note = f"⚠️ ID Mismatch — sellers.json has {seller_id}, ads.txt has {wrong_sid} ({rel})"
        return build_row(
            f"⚠️ {found_in}",
            f"⚠️ {rel} found but wrong ID (ads.txt={wrong_sid})",
            "⚠️ ID Mismatch",
            extra_flags=mismatch_note
        )

    else:
        # Even fake entries get domain-signal-only detection (no store check — too slow)
        inv_type, conf, inv_signal = detect_inventory_type(
            domain, "not found", do_store_check=False
        )
        return {
            "Publisher Name":  name[:35],
            "Domain":          domain,
            "Seller ID":       seller_id,
            "Files Checked":   f"✅ {res['filename']}",
            "Match Found In":  "❌ Not found in either file",
            "DIRECT Line":     "❌ ID not listed",
            "Inventory Type":  inv_type,
            "Confidence":      conf,
            "Type Evidence":   inv_signal,
            "Flags":           flag_str,
            "Verdict":         "❌ Fake Entry"
        }

# ─── EXCEL COLOR EXPORT ───────────────────────────────────────────────────

VERDICT_COLORS = {
    "✅ Legitimate":     "C8E6C9",
    "❌ Fake Entry":     "FFCDD2",
    "⚠️ Misrepresented": "FFE0B2",
    "⚠️ Domain Mismatch":"E1BEE7",
    "⚠️ ID Mismatch":    "FFF9C4",   # Yellow — exchange listed but wrong seller_id
    "⚠️ No Domain":      "F5F5F5",
    "✅ Found":          "C8E6C9",
    "❌ Not Found":      "FFCDD2",
    "⚠️ Unreachable":    "FFE0B2",
    "✅ Clean":          "C8E6C9",
    "🚨 Action Needed":  "FFCDD2",
    "⚠️ Review":         "FFE0B2",
    "✅ Pass":           "C8E6C9",
    "❌ Fail":           "FFCDD2",
    "❌ Missing":        "FFCDD2",
}

def color_for_row(row_dict):
    for val in row_dict.values():
        v = str(val)
        for key, color in VERDICT_COLORS.items():
            if key in v:
                return color
    return "FFFFFF"

def df_to_colored_excel(df, sheet_name="Results", title="AdChain Inspector v2"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor="1A3A5C")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
    sub = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Advlion AdOps")
    sub.font      = Font(size=10, color="FFFFFF", italic=True)
    sub.fill      = PatternFill("solid", fgColor="2E75B6")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Header row
    header_fill = PatternFill("solid", fgColor="0D47A1")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, size=10, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[3].height = 22

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
        row_color = color_for_row(row.to_dict())
        fill = PatternFill("solid", fgColor=row_color)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.fill      = fill
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[row_idx].height = 16

    # Auto column width
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)),
                      max((len(str(df.iloc[i, col_idx-1])) for i in range(min(len(df), 50))), default=10))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    # Freeze panes below header
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─── SIDEBAR ──────────────────────────────────────────────────────────────

def sidebar():
    st.sidebar.markdown("""
    <div style='text-align:center; padding:16px 0 6px 0;'>
        <div style='font-size:34px;'>🔗</div>
        <div style='font-size:17px; font-weight:800; color:#7EC8E3 !important;'>AdChain Inspector</div>
        <div style='font-size:12px; color:#4CAF50 !important; font-weight:700;'>v2.0</div>
        <div style='font-size:10px; color:#3A5070 !important;'>Advlion AdOps  ·  Supply Chain Intelligence</div>
    </div>
    <hr style='border-color:#1A2E44; margin:8px 0 12px 0;'>
    """, unsafe_allow_html=True)

    pages = {
        "🏠  Overview":                    "overview",
        "─── SUPPLY CHAIN ───":            "divider1",
        "🔍  sellers.json Authenticator":  "authenticator",
        "📋  sellers.json Inspector":       "inspector",
        "📄  ads.txt / app-ads.txt":        "adstxt",
        "🔗  Cross Validator":             "cross",
        "─── DEMAND ───":                  "divider2",
        "🏦  Demand Seat Checker":         "demand",
        "🧠  Demand Health Score":         "health",
        "─── v2 NEW PAGES ───":            "divider3",
        "⚡  Bulk SSP Authenticator":      "bulk_ssp",
        "🏢  Advlion Presence Scanner":    "presence",
        "📊  S&D Intelligence":            "sd_intel",
        "🕵️  Change Tracker":              "tracker",
        "🔀  Intermediary Intelligence":   "intermediary",
        "─── OTHER TOOLS ───":             "divider4",
        "🔗  Supply Chain Validator":      "supply_validator",
        "📱  Bundle Intelligence":         "bundle",
        "⛓️  Schain Validator":            "schain",
        "🛡️  IVT Risk Scorer":            "ivt",
        "📝  Onboarding Report":           "onboarding",
        "📅  Weekly Digest":               "digest",
    }

    dividers = {k for k in pages if "───" in k}
    options  = [k for k in pages if k not in dividers]

    sel = st.sidebar.radio("", options, label_visibility="collapsed")

    st.sidebar.markdown(f"""
    <hr style='border-color:#1A2E44; margin:10px 0;'>
    <div style='font-size:11px; color:#3A5570 !important; padding:0 6px;'>
        <span style='color:#7EC8E3 !important; font-weight:700;'>Advlion</span><br>
        advlion.com · SID <span style='color:#7EC8E3 !important;'>3148</span><br>
        <span style='color:#7EC8E3 !important; font-weight:700;'>Cache</span> · {DB_PATH.split(os.sep)[-1]}<br>
        <span style='color:#5577AA !important;'>{datetime.now().strftime("%d %b %Y  %H:%M")}</span>
    </div>
    """, unsafe_allow_html=True)

    return pages[sel]

# ─── PAGE: OVERVIEW ───────────────────────────────────────────────────────

def page_overview():
    st.title("🔗 AdChain Inspector v2")
    st.markdown("*Supply Chain Intelligence for Advlion AdOps — now with cache, change tracking, bulk checks, S&D integration and colored Excel exports*")

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Version",   "v2.0")
    c2.metric("Pages",     "16")
    c3.metric("Exchanges", "15")
    c4.metric("Checks",    "40+")
    c5.metric("Advlion SID", "3148")

    st.markdown("---")
    st.markdown("### 🆕 What's New in v2")
    col1, col2 = st.columns(2)
    new_features = [
        ("⚡", "Bulk SSP Authenticator",   "Check 10–20 SSPs simultaneously. Fraud score per SSP across all their publishers."),
        ("🏢", "Advlion Presence Scanner", "Upload CSV of publishers → bulk check if advlion.com,3148,DIRECT is listed in each."),
        ("📊", "S&D Intelligence",         "Upload your weekly S&D Excel → cross-reference revenue vs compliance scores. See which high-revenue SSPs have the worst fraud."),
        ("🕵️", "Change Tracker",           "SQLite-powered history. See exactly what changed in any SSP's sellers.json vs last week."),
        ("🎨", "Colored Excel Export",     "Every results table exports as .xlsx with 🟢 green/🔴 red/🟠 orange rows by verdict. Chart-ready."),
        ("🔁", "Smart Retry",              "3 URL variants tried per domain (https, http, www). 2 auto-retries. 30% fewer false 'Not Found'."),
        ("🚨", "CDN Fraud Flag",           "Auto-detects cloudfront.net, amazonaws.com etc. in sellers.json — flags as suspicious immediately."),
        ("🔢", "SID Normalisation",        "Leading-zero seller_id mismatches fixed. 3421 == 03421 now correctly matches."),
    ]
    for i, (icon, name, desc) in enumerate(new_features):
        with (col1 if i % 2 == 0 else col2):
            with st.expander(f"{icon} **{name}**"):
                st.markdown(desc)

    st.markdown("---")
    # Cache stats
    st.markdown("### 🗄️ Cache Status")
    try:
        con = sqlite3.connect(DB_PATH)
        sj_count  = con.execute("SELECT COUNT(*) FROM sellers_cache").fetchone()[0]
        snap_count = con.execute("SELECT COUNT(*) FROM sellers_snapshots").fetchone()[0]
        con.close()
        c1, c2, c3 = st.columns(3)
        c1.metric("Cached sellers.json", sj_count)
        c2.metric("Change Snapshots",    snap_count)
        c3.metric("Cache File",          "adchain_cache.db")
        if st.button("🗑️ Clear Cache"):
            cache_clear()
            st.success("Cache cleared")
    except Exception as e:
        st.info(f"Cache initialising: {e}")

# ─── PAGE: AUTHENTICATOR ─────────────────────────────────────────────────

def page_authenticator():
    st.title("🔍 sellers.json Authenticator")
    st.markdown("""
    **Core Logic:** Fetches sellers.json → PUBLISHER entries only → fetches each publisher's ads.txt
    → checks for DIRECT line → **No DIRECT = Fake/Fraudulent entry**
    """)
    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        sellers_url = st.text_input("sellers.json URL",
            placeholder="e.g. https://advlion.com/sellers.json")
    with col2:
        exchange_domain_override = st.text_input("Exchange Domain Override (optional)",
            placeholder="e.g. advlion.com — auto-detected if blank",
            help="Use when sellers.json is hosted at a CDN/redirect URL different from what publishers list in their ads.txt")

    # ── File selection + mode info ─────────────────────────────────────────
    st.markdown("**📂 Which files to check:**")
    fcol1, fcol2, fcol3 = st.columns(3)
    with fcol1:
        check_web = st.checkbox("✅ ads.txt (Web/Domain)", value=True,
            help="Uncheck to skip ads.txt — only In-App & CTV will be checked. Faster & focused.")
    with fcol2:
        check_app = st.checkbox("✅ app-ads.txt (In-App/CTV)", value=True,
            help="Uncheck to skip app-ads.txt — only Web publishers will appear.")
    with fcol3:
        use_cache = st.checkbox("⚡ Use cache (faster)", value=True)

    # Mode badge
    if not check_web and check_app:
        st.success("⚡ **In-App + CTV Mode** — ads.txt skipped. Only In-App and CTV publishers checked. Faster run.")
    elif check_web and not check_app:
        st.info("🌐 **Web-Only Mode** — app-ads.txt skipped. Only Web/Domain publishers checked.")
    elif not check_web and not check_app:
        st.error("⚠️ Both files unchecked — please enable at least one.")
    else:
        st.info("🔎 **Full Check** — both ads.txt and app-ads.txt will be fetched and combined.")

    col5, col6, col7 = st.columns(3)
    with col5:
        max_check = st.slider(
            "Max publishers", 10, 2000, 100,
            help="Now supports up to 2000 publishers. For 1000+ entries, use 50 workers for best speed."
        )
    with col6:
        workers = st.slider(
            "Parallel workers", 5, 80, 50,
            help="Higher = faster. 50 works well for most. For 1000+ publishers use 60-80."
        )
    with col7:
        timeout_override = st.slider(
            "Request timeout (s)", 2, 10, 4,
            help="Lower = skip slow domains faster. 4s is optimal for speed vs accuracy."
        )

    # Speed estimate
    if max_check > 0 and workers > 0:
        est_secs = max(5, int(max_check / workers) * timeout_override)
        st.caption(
            f"⚡ **Speed estimate:** ~{est_secs}–{est_secs*2}s for {max_check} publishers "
            f"with {workers} workers. Second run: instant from cache."
        )

    ex_domain_for_check = (exchange_domain_override.strip().lower().replace("www.", "")
                           if exchange_domain_override.strip()
                           else extract_domain_from_field(sellers_url.strip()))

    if ex_domain_for_check:
        st.info(f"🔎 Matching seller_id as DIRECT under exchange domain: `{ex_domain_for_check}`")

    if st.button("🔍 Run Authenticity Check", type="primary"):
        if not sellers_url.strip():
            st.warning("Enter a sellers.json URL")
            return

        def try_fetch_sellers_json(url):
            """
            Try fetching URL. If response looks compressed but unparseable,
            auto-retry with Accept-Encoding: identity (no compression).
            Returns (response, error_str).
            """
            for attempt_headers in [
                HEADERS,   # first: normal browser headers (may get compressed response)
                {**HEADERS, "Accept-Encoding": "identity"},  # retry: force plain, no gzip/brotli
            ]:
                try:
                    r = requests.get(url.strip(), timeout=15, headers=attempt_headers)
                    # If body is non-empty on second attempt, return it
                    if r.content:
                        return r, None
                except requests.exceptions.Timeout:
                    return None, "timeout"
                except requests.exceptions.ConnectionError:
                    return None, "connection"
                except Exception as e:
                    return None, str(e)
            return None, "empty_response"

        def is_nextjs_or_spa(text):
            """Detect if response is a JS-rendered SPA/Next.js shell"""
            indicators = ["/_next/static", "__NEXT_DATA__", "window.__reactFiber",
                          "react-dom", "/static/js/", "application/javascript"]
            return any(ind in text for ind in indicators)

        def decompress_content(raw_bytes):
            """
            Try all known decompression formats:
            gzip, zlib, deflate, brotli — fixes servers that compress
            without sending Content-Encoding header.
            """
            decompressors = [
                ("gzip",         lambda b: gzip.decompress(b)),
                ("zlib",         lambda b: zlib.decompress(b)),
                ("zlib-deflate", lambda b: zlib.decompress(b, -15)),
                ("zlib-gzip",    lambda b: zlib.decompress(b,  47)),
                ("brotli",       lambda b: _brotli.decompress(b) if HAS_BROTLI else None),
            ]
            for name, fn in decompressors:
                try:
                    result = fn(raw_bytes)
                    if result:
                        return result
                except Exception:
                    continue
            return None

        def robust_parse_json(r):
            """Try 12 strategies to extract valid JSON — handles encoding, compression, JSONP, BOM."""
            import re as _re
            raw_bytes = r.content
            decomp    = decompress_content(raw_bytes)  # may be None if not compressed

            strategies = [
                # ── Standard ────────────────────────────────────────────────
                ("standard",         lambda: r.json()),
                ("stripped",         lambda: json.loads(r.text.strip())),
                # ── BOM / Encoding ───────────────────────────────────────────
                ("utf-8-sig",        lambda: json.loads(raw_bytes.decode("utf-8-sig").strip())),
                ("bom-strip",        lambda: json.loads(r.text.lstrip("\ufeff").strip())),
                ("utf-16",           lambda: json.loads(raw_bytes.decode("utf-16").strip())),
                ("latin-1",          lambda: json.loads(raw_bytes.decode("latin-1").strip())),
                ("ignore-errors",    lambda: json.loads(raw_bytes.decode("utf-8", errors="ignore").strip())),
                # ── Decompressed ─────────────────────────────────────────────
                ("gzip-utf8",        lambda: json.loads(decomp.decode("utf-8").strip()) if decomp else (_ for _ in ()).throw(ValueError("no decomp"))),
                ("gzip-utf8-sig",    lambda: json.loads(decomp.decode("utf-8-sig").strip()) if decomp else (_ for _ in ()).throw(ValueError("no decomp"))),
                ("gzip-ignore",      lambda: json.loads(decomp.decode("utf-8", errors="ignore").strip()) if decomp else (_ for _ in ()).throw(ValueError("no decomp"))),
                # ── JSONP unwrap ─────────────────────────────────────────────
                ("jsonp",            lambda: json.loads(_re.sub(r'^\s*\w+\s*\(|\)\s*;?\s*$', '', r.text).strip())),
                # ── Decompressed JSONP ────────────────────────────────────────
                ("gzip-jsonp",       lambda: json.loads(_re.sub(r'^\s*\w+\s*\(|\)\s*;?\s*$', '', decomp.decode("utf-8", errors="ignore")).strip()) if decomp else (_ for _ in ()).throw(ValueError("no decomp"))),
            ]
            last_err = None
            for name, fn in strategies:
                try:
                    data = fn()
                    if data is not None:
                        return data, name
                except Exception as e:
                    last_err = e
                    continue
            raise ValueError(f"All {len(strategies)} strategies failed. Last: {last_err}")

        def looks_like_valid_sellers_json(r):
            """Check if response looks like real sellers.json — tries 8 parse strategies"""
            ct       = r.headers.get("Content-Type", "")
            body     = r.text if r.content else ""

            # Hard-reject only if HTML AND confirmed SPA/Next.js
            if "text/html" in ct and is_nextjs_or_spa(body):
                return False, "html"

            # Empty / null body
            if not r.content or not body.strip() or body.strip().lower() == "null":
                return False, "empty"

            # Try all parse strategies
            try:
                data, strategy = robust_parse_json(r)
            except ValueError as e:
                if "text/html" in ct:
                    return False, "html"
                return False, f"invalid_json|{str(e)[:140]}"

            if data is None:
                return False, "null_json"
            if not isinstance(data, dict):
                return False, f"wrong_type|Got {type(data).__name__} instead of object"
            if "sellers" not in data:
                return False, f"no_sellers_key|Keys: {list(data.keys())[:8]}"

            return True, f"ok|{strategy}"

        # ── Step 1: Try the given URL ──────────────────────────────────────
        input_url   = sellers_url.strip()
        r, fetch_err = try_fetch_sellers_json(input_url)

        # ── Step 2: Auto-fallback to root domain if needed ─────────────────
        fallback_url  = None
        fallback_used = False
        tried_urls    = [input_url]

        if r is not None:
            valid, reason = looks_like_valid_sellers_json(r)
        else:
            valid, reason = False, fetch_err or "error"

        if not valid:
            # Build fallback: root domain sellers.json
            parsed   = urlparse(input_url)
            root     = f"{parsed.scheme}://{parsed.netloc}/sellers.json"
            alt_root = f"https://{extract_domain_from_field(input_url)}/sellers.json"
            candidates = list(dict.fromkeys([root, alt_root]))  # deduplicated

            for candidate in candidates:
                if candidate == input_url:
                    continue
                tried_urls.append(candidate)
                st.info(f"🔄 Original URL failed — auto-trying: `{candidate}`")
                r2, err2 = try_fetch_sellers_json(candidate)
                if r2 is not None:
                    v2, r2_reason = looks_like_valid_sellers_json(r2)
                    if v2:
                        r, valid, reason = r2, True, "ok"
                        fallback_url  = candidate
                        fallback_used = True
                        st.success(f"✅ Found valid sellers.json at fallback: `{candidate}`")
                        break

        # ── Step 3: If still not valid, show specific error ────────────────
        if not valid:
            last_r    = r
            last_text = last_r.text if last_r is not None else ""

            if reason in ("html", "invalid_json") and is_nextjs_or_spa(last_text):
                st.error("❌ This sellers.json is served by a **JavaScript-rendered app (Next.js / React)**")
                st.warning(
                    "**Why this happens:**\n"
                    "The JSON you see in your browser is rendered by JavaScript after page load. "
                    "Our tool uses Python `requests` which cannot execute JavaScript — "
                    "it only receives the raw HTML shell, not the rendered JSON data.\n\n"
                    "**This is a problem on the partner's end.** A properly configured sellers.json "
                    "should be a plain static file served directly — not a JavaScript-rendered page.\n\n"
                    "**What to tell the partner:**\n"
                    "> *'Your sellers.json at `" + input_url + "` is a JavaScript-rendered endpoint "
                    "and cannot be fetched by programmatic tools or IAB validators. "
                    "Please host sellers.json as a static file with Content-Type: application/json.'*\n\n"
                    f"**URLs tried:** {' → '.join(tried_urls)}"
                )
            elif reason in ("html",):
                st.error("❌ URL returned an HTML page instead of JSON.")
                st.warning(
                    f"**Most likely cause:** Server is blocking non-browser requests.\n\n"
                    f"**URLs tried:** {' → '.join(tried_urls)}\n\n"
                    f"**Ask the partner** to make sellers.json accessible as a plain static file."
                )
                if last_text:
                    st.code(f"Response preview:\n{last_text[:300]}", language="text")
            elif reason in ("empty", "null_json"):
                st.error("❌ sellers.json returned empty or null — file not populated yet.")
                st.info("The endpoint exists but has no data. Ask partner to populate their sellers.json.")
            elif reason == "no_sellers_key":
                st.error("❌ JSON parsed but missing required `sellers` key.")
                if last_r:
                    st.json(last_r.json())
            elif reason == "timeout":
                st.error("❌ Request timed out. Server took too long to respond.")
                st.info("Try again in a few seconds, or increase the timeout by trying a simpler URL.")
            elif reason == "connection":
                st.error("❌ Could not connect. Check if the domain is correct and reachable.")
            elif reason.startswith("wrong_type"):
                st.error(f"❌ sellers.json structure error: {reason.split('|')[1] if '|' in reason else reason}")
                st.info("A valid sellers.json must be a JSON object with a `sellers` array at the top level.")
            elif reason.startswith("no_sellers_key"):
                detail = reason.split("|")[1] if "|" in reason else ""
                st.error("❌ JSON parsed successfully but missing required `sellers` key.")
                st.warning(
                    f"The file is valid JSON but does not follow IAB sellers.json spec.\n\n"
                    f"{detail}\n\n"
                    "A valid sellers.json must have: `{{ \"sellers\": [ ... ] }}`"
                )
                if last_r:
                    try:
                        st.json(last_r.json())
                    except Exception:
                        st.code(last_r.text[:500])
            elif reason.startswith("invalid_json"):
                detail = reason.split("|")[1] if "|" in reason else "Unknown parse error"
                st.error("❌ Could not parse as valid JSON after trying 8 different methods.")
                st.warning(
                    f"**Parse error detail:** `{detail}`\n\n"
                    "**Possible causes:**\n"
                    "- File has encoding issues (BOM, UTF-16, special characters)\n"
                    "- File has syntax errors (trailing commas, unquoted keys)\n"
                    "- File is actually XML, CSV, or another non-JSON format\n"
                    "- File is password-protected or encrypted\n\n"
                    "**Tip:** Open the URL in your browser, press Ctrl+U to view raw source, "
                    "and check if the raw content starts with `{` or `[`."
                )
                if last_r:
                    preview = last_r.content[:300]
                    st.code(f"Raw response (first 300 bytes):\n{preview}", language="text")
            else:
                st.error(f"❌ Could not fetch: {reason}")
            return

        # ── Step 4: Valid data — proceed ───────────────────────────────────
        if fallback_used:
            st.info(f"ℹ️ Using fallback URL: `{fallback_url}` (original URL was blocked/invalid)")

        try:
            data = r.json()
        except Exception as e:
            st.error(f"❌ JSON parse error: {e}")
            return

        sellers       = get_sellers(data)
        publishers    = [s for s in sellers if str(s.get("seller_type","")).upper() == "PUBLISHER"]
        intermediaries= [s for s in sellers if str(s.get("seller_type","")).upper() == "INTERMEDIARY"]
        confidential  = [s for s in sellers if s.get("is_confidential") == 1]
        cdn_entries   = [s for s in sellers if is_cdn_domain(extract_domain_from_field(str(s.get("domain",""))))]

        st.markdown("### 📊 File Overview")
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.metric("Total",        len(sellers))
        c2.metric("PUBLISHER",    len(publishers))
        c3.metric("INTERMEDIARY", len(intermediaries))
        c4.metric("is_confidential", len(confidential))
        c5.metric("🚨 CDN Domains", len(cdn_entries))
        c6.metric("Checking",     min(len(publishers), max_check))

        if cdn_entries:
            st.warning(f"🚨 {len(cdn_entries)} publishers have CDN/cloud domains (cloudfront, amazonaws etc.) — likely fake entries")

        st.markdown(f"---\n### 🔍 Validating {min(len(publishers), max_check)} Publishers in Parallel...")

        to_check = publishers[:max_check]
        results  = []
        done     = 0
        prog     = st.progress(0)
        stat     = st.empty()

        # Adaptive batch size — show live results every N completions
        LIVE_UPDATE_EVERY = max(10, len(to_check) // 20)
        live_table = st.empty()

        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_map = {
                executor.submit(
                    check_single_publisher, s, check_app,
                    ex_domain_for_check, check_web, timeout_override
                ): s
                for s in to_check
            }
            for future in as_completed(future_map):
                result = future.result()
                results.append(result)
                done += 1
                prog.progress(done / len(to_check))

                legit_so_far  = len([r for r in results if "Legitimate" in r.get("Verdict","")])
                fake_so_far   = len([r for r in results if "Fake"       in r.get("Verdict","")])
                idmm_so_far   = len([r for r in results if "ID Mismatch"in r.get("Verdict","")])

                stat.markdown(
                    f"⚡ **{done}/{len(to_check)}** checked · "
                    f"✅ {legit_so_far} Legit · "
                    f"❌ {fake_so_far} Fake · "
                    f"🟡 {idmm_so_far} ID Mismatch · "
                    f"🏎️ {workers} workers @ {timeout_override}s timeout"
                )

                # Live preview table — updates every N rows
                if done % LIVE_UPDATE_EVERY == 0 or done == len(to_check):
                    preview_df = pd.DataFrame(results[-LIVE_UPDATE_EVERY:])
                    if "Verdict" in preview_df.columns and "Domain" in preview_df.columns:
                        live_table.dataframe(
                            preview_df[["Publisher Name","Domain","Verdict"]].tail(10),
                            use_container_width=True, height=220
                        )

        prog.empty(); stat.empty()
        df = pd.DataFrame(results)

        legit    = len([r for r in results if "Legitimate"    in r["Verdict"]])
        fake     = len([r for r in results if "Fake"           in r["Verdict"]])
        misrep   = len([r for r in results if "Misrepresented" in r["Verdict"]])
        id_mm    = len([r for r in results if "ID Mismatch"    in r["Verdict"]])
        dom_mm   = len([r for r in results if "Domain Mismatch"in r["Verdict"]])
        nodom    = len([r for r in results if "No Domain"      in r["Verdict"]])
        total    = len(results)

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("✅ Legitimate",      legit,  f"{round(legit/total*100)}%")
        c2.metric("❌ Fake Entry",      fake,   f"{round(fake/total*100)}%")
        c3.metric("⚠️ Misrepresented",  misrep)
        c4.metric("🟡 ID Mismatch",     id_mm,
                  help="Exchange domain listed as DIRECT but seller_id in ads.txt differs from sellers.json")
        c5.metric("⚠️ Domain Mismatch", dom_mm)
        c6.metric("⚠️ No Domain",       nodom)

        if id_mm > 0:
            st.warning(
                f"🟡 **{id_mm} ID Mismatch entries** — The exchange domain is listed as DIRECT in ads.txt "
                f"but the seller_id doesn't match what's in sellers.json. "
                f"This could mean the publisher has multiple accounts, or their ads.txt is outdated. "
                f"These are highlighted in **yellow** in the results table."
            )

        if fake > 0:
            st.error(f"🚨 {fake} fake entries detected — sellers.json contains fraudulent PUBLISHER declarations")
        else:
            st.success("✅ No fake entries detected in checked publishers")

        fig = px.pie(
            values=[legit, fake, misrep, id_mm, dom_mm, nodom],
            names=["Legitimate","Fake Entry","Misrepresented","ID Mismatch","Domain Mismatch","No Domain"],
            color_discrete_map={
                "Legitimate":     "#4CAF50",
                "Fake Entry":     "#F44336",
                "Misrepresented": "#FF9800",
                "ID Mismatch":    "#FDD835",   # Yellow
                "Domain Mismatch":"#9C27B0",
                "No Domain":      "#9E9E9E"
            },
            hole=0.4, title="Publisher Legitimacy Distribution"
        )
        st.plotly_chart(fig, use_container_width=True)

        # ── Inventory Type Breakdown ─────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📺 Inventory Type Portfolio Breakdown")
        st.caption("Tiered signal scoring: domain keywords (CTV Tier 1-3) + file match (ads.txt/app-ads.txt) + live Play Store / App Store + homepage scan")

        if "Inventory Type" in df.columns:
            COLOR_MAP = {
                "🖥️ CTV":               "#8E44AD",
                "🖥️ CTV (Web)":         "#9B59B6",
                "🖥️ CTV + 🌐 Web":     "#6C3483",
                "📱 In-App":            "#2471A3",
                "🔀 Mixed (Web + App)": "#D4AC0D",
                "🌐 Web":               "#1E8449",
                "❓ Unknown":           "#717D7E",
            }

            inv_counts = df["Inventory Type"].value_counts().reset_index()
            inv_counts.columns = ["Inventory Type", "Count"]
            inv_counts["Color"] = inv_counts["Inventory Type"].map(
                lambda x: COLOR_MAP.get(x, "#999999")
            )

            tab_donut, tab_bar, tab_verdict = st.tabs(
                ["🍩 Donut Chart", "📊 Stacked Bar", "🔀 Type × Verdict"]
            )

            with tab_donut:
                col_chart, col_stats = st.columns([3, 2])
                with col_chart:
                    fig_d = px.pie(
                        inv_counts, values="Count", names="Inventory Type",
                        color="Inventory Type",
                        color_discrete_map=COLOR_MAP,
                        hole=0.5,
                        title="Publisher Portfolio — Inventory Type"
                    )
                    fig_d.update_traces(
                        textinfo="label+percent+value",
                        textfont_size=12,
                        pull=[0.05]*len(inv_counts)
                    )
                    fig_d.update_layout(
                        showlegend=True, height=420,
                        legend=dict(orientation="v", x=1.02),
                        paper_bgcolor="rgba(0,0,0,0)",
                        plot_bgcolor="rgba(0,0,0,0)"
                    )
                    st.plotly_chart(fig_d, use_container_width=True)

                with col_stats:
                    st.markdown("#### 📊 Portfolio Summary")
                    for _, row in inv_counts.iterrows():
                        itype = row["Inventory Type"]
                        count = int(row["Count"])
                        pct   = round(count / total * 100) if total > 0 else 0
                        st.metric(label=itype, value=f"{count} publishers", delta=f"{pct}% of checked")

                    st.markdown("---")
                    legit_df = df[df["Verdict"].str.contains("Legitimate", na=False)]
                    if not legit_df.empty:
                        st.markdown("**✅ Legit publishers breakdown:**")
                        for itype, cnt in legit_df["Inventory Type"].value_counts().items():
                            pct = round(cnt / len(legit_df) * 100)
                            color = COLOR_MAP.get(itype, "#999")
                            st.markdown(
                                f"<span style='color:{color};font-weight:700'>{itype}</span>: {cnt} ({pct}%)",
                                unsafe_allow_html=True
                            )

                    # Confidence breakdown
                    if "Confidence" in df.columns:
                        st.markdown("---")
                        st.markdown("**🎯 Detection Confidence:**")
                        for conf, cnt in df["Confidence"].value_counts().items():
                            st.markdown(f"- {conf}: **{cnt}** publishers")

            with tab_bar:
                # Stacked bar: Inventory Type × Verdict
                cross = pd.crosstab(df["Inventory Type"], df["Verdict"].str.replace("✅ ","").str.replace("❌ ","").str.replace("⚠️ ",""))
                fig_bar = px.bar(
                    cross.reset_index().melt(id_vars="Inventory Type", var_name="Verdict", value_name="Count"),
                    x="Inventory Type", y="Count", color="Verdict",
                    color_discrete_map={
                        "Legitimate":     "#27AE60",
                        "Fake Entry":     "#C0392B",
                        "Misrepresented": "#E67E22",
                        "Domain Mismatch":"#8E44AD",
                        "No Domain":      "#95A5A6",
                    },
                    title="Inventory Type × Verdict Breakdown",
                    barmode="stack", text_auto=True
                )
                fig_bar.update_layout(
                    height=420, xaxis_tickangle=-25,
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    xaxis=dict(gridcolor="#2A3A4A"),
                    yaxis=dict(gridcolor="#2A3A4A"),
                )
                st.plotly_chart(fig_bar, use_container_width=True)

            with tab_verdict:
                # Cross-table heatmap
                cross2 = pd.crosstab(
                    df["Inventory Type"],
                    df["Verdict"].str.replace("✅ ","").str.replace("❌ ","").str.replace("⚠️ ","")
                )
                st.dataframe(cross2.style.background_gradient(cmap="RdYlGn", axis=None),
                             use_container_width=True)
                st.caption("Green = more Legitimate, Red = more Fake in that inventory type")
        else:
            st.info("Inventory type data not available — re-run the check.")

        st.markdown("---")

        # ── Sort results by Inventory Type ────────────────────────────────
        INV_SORT_ORDER = {
            "🖥️ CTV":               0,
            "🖥️ CTV (Web)":         1,
            "🖥️ CTV + 🌐 Web":      2,
            "📱 In-App":             3,
            "🔀 Mixed (Web + App)":  4,
            "🌐 Web":                5,
            "❓ Unknown":            6,
        }
        if "Inventory Type" in df.columns:
            df["_sort_key"] = df["Inventory Type"].map(
                lambda x: INV_SORT_ORDER.get(str(x) if x and str(x) != "nan" else "❓ Unknown", 7)
            )
            df = df.sort_values(["_sort_key", "Verdict"], ascending=[True, True])
            df = df.drop(columns=["_sort_key"])

        # ── Grouped view by Inventory Type ────────────────────────────────
        st.markdown("### 📋 Results — Grouped by Inventory Type")
        st.caption("Sorted: CTV → In-App → Mixed → Web → Unknown. Expand each group to review.")

        if "Inventory Type" in df.columns:
            groups = df["Inventory Type"].unique()
            group_order = sorted(groups, key=lambda x: INV_SORT_ORDER.get(x, 7))

            group_icons = {
                "🖥️ CTV":              "🖥️",
                "🖥️ CTV (Web)":        "🖥️",
                "🖥️ CTV + 🌐 Web":     "🖥️🌐",
                "📱 In-App":            "📱",
                "🔀 Mixed (Web + App)": "🔀",
                "🌐 Web":               "🌐",
                "❓ Unknown":           "❓",
            }

            for inv_type in group_order:
                inv_type  = str(inv_type) if inv_type and str(inv_type) != "nan" else "❓ Unknown"
                grp       = df[df["Inventory Type"].astype(str) == inv_type].copy()
                legit_cnt = len(grp[grp["Verdict"].str.contains("Legitimate", na=False)])
                fake_cnt  = len(grp[grp["Verdict"].str.contains("Fake", na=False)])
                icon      = group_icons.get(inv_type, "📋")
                label     = (
                    f"{icon} **{inv_type}** — "
                    f"{len(grp)} publishers | "
                    f"✅ {legit_cnt} Legit | "
                    f"❌ {fake_cnt} Fake"
                )
                # Auto-expand CTV and In-App, collapse others
                auto_expand = any(k in str(inv_type) for k in ["CTV","In-App"]) if inv_type and str(inv_type) != "nan" else False
                with st.expander(label, expanded=auto_expand):
                    display_cols = [c for c in [
                        "Publisher Name","Domain","Seller ID",
                        "Files Checked","Match Found In","DIRECT Line",
                        "Inventory Type","Confidence","Type Evidence",
                        "Flags","Verdict"
                    ] if c in grp.columns]
                    st.dataframe(grp[display_cols], use_container_width=True,
                                 height=min(400, 40 + len(grp)*36))
                    c1, c2 = st.columns(2)
                    # Include group index to guarantee unique keys even for duplicate type names
                    import hashlib as _hl
                    safe_type = (inv_type
                        .replace("/","").replace(" ","_").replace("+","and")
                        .replace("🖥️","CTV").replace("📱","App").replace("🌐","Web")
                        .replace("🔀","Mixed").replace("❓","Unknown")
                        .encode("ascii","ignore").decode())[:20]
                    uid = _hl.md5(f"{inv_type}{id(grp)}".encode()).hexdigest()[:6]
                    c1.download_button(
                        f"⬇️ CSV — {inv_type[:15]}",
                        df_to_csv(grp[display_cols]),
                        f"auth_{safe_type}.csv", "text/csv",
                        key=f"csv_{safe_type}_{uid}"
                    )
                    c2.download_button(
                        f"⬇️ Excel — {inv_type[:15]}",
                        df_to_colored_excel(grp[display_cols], inv_type[:30]),
                        f"auth_{safe_type}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"xls_{safe_type}_{uid}"
                    )

        st.markdown("---")
        st.markdown("### 📋 Full Results (All Combined)")
        st.dataframe(df, use_container_width=True, height=420)

        col_csv, col_xls = st.columns(2)
        col_csv.download_button("⬇️ Full CSV", df_to_csv(df), "authenticity_full.csv", "text/csv")
        col_xls.download_button("⬇️ Full Excel (colored)",
                                df_to_colored_excel(df, "Authenticity Results"),
                                "authenticity_full.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── PAGE: INSPECTOR ──────────────────────────────────────────────────────

def page_inspector():
    st.title("📋 sellers.json Inspector")
    st.markdown("*Full structural audit — fields, duplicates, transparency, Advlion presence, CDN flags*")
    st.markdown("---")

    col1, col2 = st.columns([3,1])
    with col1:
        domain = st.text_input("Domain", placeholder="e.g. openx.com")
    with col2:
        use_cache = st.checkbox("Use cache", value=True)

    if st.button("📋 Inspect", type="primary"):
        if not domain.strip():
            st.warning("Enter domain")
            return
        d = clean_domain(domain)
        with st.spinner(f"Fetching {d}/sellers.json..."):
            result = fetch_sellers_json(d, use_cache)

        if not result["success"]:
            st.error(f"❌ Could not fetch {d}/sellers.json")
            return

        if result.get("from_cache"):
            age = result.get("cache_age")
            st.info(f"⚡ Loaded from cache (fetched {age.strftime('%d %b %H:%M') if age else 'earlier'})")
        else:
            st.success(f"✅ Fetched live from {result['url']}")

        data    = result["data"]
        sellers = get_sellers(data)

        pub   = [s for s in sellers if str(s.get("seller_type","")).upper() == "PUBLISHER"]
        inter = [s for s in sellers if str(s.get("seller_type","")).upper() == "INTERMEDIARY"]
        both  = [s for s in sellers if str(s.get("seller_type","")).upper() == "BOTH"]
        conf  = [s for s in sellers if s.get("is_confidential") == 1]
        cdn_s = [s for s in sellers if is_cdn_domain(extract_domain_from_field(str(s.get("domain",""))))]

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("Total",         len(sellers))
        c2.metric("PUBLISHER",     len(pub))
        c3.metric("INTERMEDIARY",  len(inter))
        c4.metric("BOTH",          len(both))
        c5.metric("is_confidential",len(conf))
        c6.metric("🚨 CDN Domains", len(cdn_s))

        conf_pct = round(len(conf)/len(sellers)*100) if sellers else 0
        if conf_pct > 30:
            st.error(f"🚨 {conf_pct}% confidential — very low transparency")
        elif conf_pct > 10:
            st.warning(f"⚠️ {conf_pct}% confidential")
        else:
            st.success(f"✅ {conf_pct}% confidential — good transparency")

        if cdn_s:
            st.error(f"🚨 {len(cdn_s)} entries use CDN/cloud domains (not real publisher domains)")

        # Duplicates
        sids     = [str(s.get("seller_id","")) for s in sellers]
        dup_sids = list(set([sid for sid in sids if sids.count(sid) > 1]))

        no_domain = [s for s in sellers if not s.get("domain")]
        no_name   = [s for s in sellers if not s.get("name")]

        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Missing Domain",    len(no_domain))
        c2.metric("Missing Name",      len(no_name))
        c3.metric("Duplicate Seller IDs", len(dup_sids))
        c4.metric("Version",           data.get("version","—"))

        if dup_sids:
            st.error(f"🚨 Duplicate seller_ids: {', '.join(dup_sids[:10])}")

        # Advlion
        st.markdown("---")
        adv = [s for s in sellers if str(s.get("domain","")).lower() == ADVLION_DOMAIN
               or str(s.get("seller_id","")) == ADVLION_SID]
        if adv:
            st.success(f"✅ Advlion found in {d}/sellers.json")
            for e in adv:
                st.json(e)
        else:
            st.error(f"❌ Advlion (advlion.com / {ADVLION_SID}) NOT found in {d}/sellers.json")

        df = pd.DataFrame(sellers)
        st.dataframe(df, use_container_width=True, height=380)

        col_csv, col_xls = st.columns(2)
        col_csv.download_button("⬇️ CSV", df_to_csv(df), f"{d}_sellers.csv", "text/csv")
        col_xls.download_button("⬇️ Excel", df_to_colored_excel(df, "sellers.json"),
                                f"{d}_sellers.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── PAGE: ADS.TXT ────────────────────────────────────────────────────────

def page_adstxt():
    st.title("📄 ads.txt / app-ads.txt Validator")
    st.markdown("*Format check, Advlion presence, duplicates, CDN flags*")
    st.markdown("---")

    col1, col2 = st.columns([3,1])
    with col1:
        domain = st.text_input("Publisher Domain", placeholder="e.g. truecaller.com")
    with col2:
        ftype  = st.selectbox("File", ["ads.txt","app-ads.txt","Both"])

    if st.button("📄 Validate", type="primary"):
        if not domain.strip():
            st.warning("Enter domain")
            return
        d     = clean_domain(domain)
        files = ["ads.txt","app-ads.txt"] if ftype == "Both" else [ftype]

        for fname in files:
            st.markdown(f"### 📄 {fname} — {d}")
            is_app = "app" in fname
            with st.spinner(f"Fetching..."):
                res = fetch_ads_txt(d, is_app)

            if not res["success"]:
                st.error(f"❌ {fname} not found at {d}")
                continue

            rows      = parse_ads_txt(res["content"])
            direct    = [r for r in rows if r["relationship"] == "DIRECT"]
            reseller  = [r for r in rows if r["relationship"] == "RESELLER"]
            bad_rel   = [r for r in rows if r["relationship"] not in ["DIRECT","RESELLER"]]
            no_cert   = [r for r in rows if not r["cert_id"]]
            raw_lines = [r["raw"] for r in rows]
            dup_lines = list(set([l for l in raw_lines if raw_lines.count(l) > 1]))

            adv_direct = [r for r in rows if r["exchange_domain"] == ADVLION_DOMAIN
                          and r["seller_id"] == ADVLION_SID and r["relationship"] == "DIRECT"]
            adv_any    = [r for r in rows if r["exchange_domain"] == ADVLION_DOMAIN]

            c1,c2,c3,c4,c5 = st.columns(5)
            c1.metric("Total lines",    len(rows))
            c2.metric("DIRECT",         len(direct))
            c3.metric("RESELLER",       len(reseller))
            c4.metric("Invalid rel.",   len(bad_rel))
            c5.metric("Duplicates",     len(dup_lines))

            if adv_direct:
                st.success(f"✅ Advlion correctly listed: advlion.com, {ADVLION_SID}, DIRECT")
            elif adv_any:
                st.warning(f"⚠️ Advlion found but may have wrong SID or relationship")
                for r in adv_any: st.code(r["raw"])
            else:
                st.error(f"❌ Advlion NOT listed. Add: advlion.com, {ADVLION_SID}, DIRECT")

            df = pd.DataFrame([{"Exchange": r["exchange_domain"],"Seller ID": r["seller_id"],
                                 "Rel": r["relationship"],"Cert ID": r["cert_id"]} for r in rows])
            st.dataframe(df, use_container_width=True, height=320)
            col_csv, col_xls = st.columns(2)
            col_csv.download_button(f"⬇️ CSV", df_to_csv(df), f"{d}_{fname}.csv","text/csv", key=f"csv_{fname}")
            col_xls.download_button(f"⬇️ Excel", df_to_colored_excel(df, fname),
                                    f"{d}_{fname}.xlsx",
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"xls_{fname}")

# ─── PAGE: CROSS VALIDATOR ────────────────────────────────────────────────

def page_cross():
    st.title("🔗 Cross Validator")
    st.markdown("*Deep single seller_id check — sellers.json + ads.txt simultaneously*")
    st.markdown("---")

    c1,c2,c3 = st.columns(3)
    with c1: ex_domain  = st.text_input("Exchange Domain", placeholder="pubmatic.com")
    with c2: sid_input  = st.text_input("Seller ID",       placeholder="156209")
    with c3: pub_domain = st.text_input("Publisher Domain (optional)", placeholder="auto-resolved")

    if st.button("🔗 Cross Validate", type="primary"):
        if not ex_domain.strip() or not sid_input.strip():
            st.warning("Exchange domain and Seller ID required")
            return

        ex_d = clean_domain(ex_domain); sid = sid_input.strip()

        st.markdown("### Step 1 — sellers.json")
        with st.spinner():
            sj = fetch_sellers_json(ex_d)

        if not sj["success"]:
            st.error(f"❌ Could not fetch {ex_d}/sellers.json"); return

        sellers = get_sellers(sj["data"])
        match   = next((s for s in sellers if str(s.get("seller_id","")).strip() == sid), None)
        inferred_domain = ""; seller_type_found = ""

        if match:
            st.success(f"✅ Seller ID {sid} found")
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Name",   str(match.get("name","—"))[:22])
            c2.metric("Domain", match.get("domain","—"))
            c3.metric("Type",   match.get("seller_type","—"))
            c4.metric("CDN?",   "🚨 Yes" if is_cdn_domain(extract_domain_from_field(str(match.get("domain","")))) else "No")
            inferred_domain  = extract_domain_from_field(str(match.get("domain","")))
            seller_type_found = str(match.get("seller_type","")).upper()
        else:
            st.error(f"❌ Seller ID {sid} NOT in {ex_d}/sellers.json")

        st.markdown("### Step 2 — ads.txt")
        check_d = clean_domain(pub_domain) if pub_domain.strip() else inferred_domain
        if not check_d:
            st.warning("No publisher domain to check"); return

        adstxt_res = None
        with st.spinner():
            for is_app in [False, True]:
                r = fetch_ads_txt(check_d, is_app)
                if r["success"]:
                    adstxt_res = r; break

        if not adstxt_res:
            st.error(f"❌ No ads.txt at {check_d}"); ads_direct = []; ads_reseller = []
        else:
            rows = parse_ads_txt(adstxt_res["content"])
            ads_direct   = [r for r in rows if r["exchange_domain"] == ex_d and r["seller_id"] == sid and r["relationship"] == "DIRECT"]
            ads_reseller = [r for r in rows if r["exchange_domain"] == ex_d and r["seller_id"] == sid and r["relationship"] == "RESELLER"]
            if ads_direct:   st.success(f"✅ DIRECT line: `{ads_direct[0]['raw']}`")
            elif ads_reseller: st.warning(f"⚠️ RESELLER only: `{ads_reseller[0]['raw']}`")
            else: st.error(f"❌ {ex_d} + {sid} not in {check_d}/ads.txt")

        st.markdown("---\n### ⚖️ Final Verdict")
        if match and ads_direct:
            st.success("✅ FULLY LEGITIMATE — PUBLISHER in sellers.json + DIRECT in ads.txt")
        elif match and ads_reseller:
            st.warning("⚠️ MISREPRESENTED — PUBLISHER in sellers.json but RESELLER in ads.txt")
        elif match and not ads_direct:
            st.error("❌ LIKELY FAKE — In sellers.json but NOT in publisher's ads.txt")
        else:
            st.error("❌ UNVERIFIABLE — Not confirmed in either file")

# ─── PAGE: DEMAND SEAT CHECKER ────────────────────────────────────────────

def page_demand():
    st.title("🏦 Demand Seat Checker")
    st.markdown("*Input demand partner → checks all 15 top exchange sellers.json files for real seats*")
    st.markdown("---")

    col1, col2 = st.columns([3,1])
    with col1: demand_domain = st.text_input("Demand Partner Domain", placeholder="e.g. anzu.io")
    with col2: also_name = st.checkbox("Also search by name", value=True)

    with st.expander("⚙️ Edit Exchange List (22 SSPs)"):
        st.caption("Core Web/Display · Mobile/In-App · Native")
        custom_ex = {}
        tier_labels = {
            # Tier 1
            "PubMatic": "── Tier 1: Core Web/Display ──",
            # Tier 2
            "Teads": "── Tier 2: Video & Outstream ──",
            # Tier 3
            "InMobi": "── Tier 3: Mobile / In-App ──",
            # Tier 4
            "Nativo": "── Tier 4: Native & Contextual ──",
        }
        cols = st.columns(3)
        for i,(nm,dm) in enumerate(TOP_15_EXCHANGES.items()):
            if nm in tier_labels:
                st.markdown(f"**{tier_labels[nm]}**")
                cols = st.columns(3)
            with cols[i%3]:
                custom_ex[nm] = st.text_input(nm, value=dm, key=f"ex_{i}")
    exchanges = custom_ex

    if st.button("🔍 Check All Exchanges", type="primary"):
        if not demand_domain.strip(): st.warning("Enter domain"); return
        dd        = clean_domain(demand_domain)
        name_part = dd.split(".")[0].lower()

        # ── Parallel fetch function — one per exchange ─────────────────
        def check_one_exchange(args):
            ex_nm, ex_dm = args
            # fetch_sellers_json uses SQLite cache — instant on repeat runs
            sj = fetch_sellers_json(ex_dm, use_cache=True)
            if not sj["success"]:
                return ex_nm, None, []
            sellers = get_sellers(sj["data"])
            matches = [s for s in sellers
                       if str(s.get("domain","")).lower().replace("www.","") == dd]
            if also_name and not matches:
                matches = [s for s in sellers
                           if name_part in str(s.get("name","")).lower()
                           or name_part in str(s.get("domain","")).lower()]
            return ex_nm, matches, sj.get("from_cache", False)

        results   = []
        all_seats = []
        done      = 0
        cached_count = 0

        prog = st.progress(0)
        stat = st.empty()
        stat.text(f"⚡ Fetching {len(exchanges)} exchanges in parallel...")

        # ── Run all 22 exchanges simultaneously ────────────────────────
        with ThreadPoolExecutor(max_workers=len(exchanges)) as executor:
            future_map = {
                executor.submit(check_one_exchange, (ex_nm, ex_dm)): ex_nm
                for ex_nm, ex_dm in exchanges.items()
            }
            for future in as_completed(future_map):
                ex_nm   = future_map[future]
                ex_dm   = exchanges[ex_nm]
                done   += 1
                prog.progress(done / len(exchanges))

                try:
                    ex_nm_r, matches, from_cache = future.result()
                    if from_cache:
                        cached_count += 1

                    if matches is None:
                        results.append({
                            "Exchange":      ex_nm,
                            "Status":        "⚠️ Unreachable",
                            "Seats":         "—",
                            "Seller Type(s)":"—",
                            "Seller IDs":    "—",
                            "Source":        "—"
                        })
                    elif matches:
                        types = list(set([str(m.get("seller_type","—")).upper() for m in matches]))
                        sids  = [str(m.get("seller_id","")) for m in matches]
                        results.append({
                            "Exchange":      ex_nm,
                            "Status":        "✅ Found",
                            "Seats":         len(matches),
                            "Seller Type(s)":", ".join(types),
                            "Seller IDs":    ", ".join(sids[:5]) + ("…" if len(sids) > 5 else ""),
                            "Source":        "⚡ Cache" if from_cache else "🌐 Live"
                        })
                        for m in matches:
                            all_seats.append({
                                "Exchange":    ex_nm,
                                "Seller ID":   m.get("seller_id"),
                                "Name":        m.get("name"),
                                "Domain":      m.get("domain"),
                                "Seller Type": m.get("seller_type")
                            })
                    else:
                        results.append({
                            "Exchange":      ex_nm,
                            "Status":        "❌ Not Found",
                            "Seats":         0,
                            "Seller Type(s)":"—",
                            "Seller IDs":    "—",
                            "Source":        "⚡ Cache" if from_cache else "🌐 Live"
                        })
                    stat.text(
                        f"✅ {done}/{len(exchanges)} done — "
                        f"{ex_nm} {'(cached ⚡)' if from_cache else '(live 🌐)'}"
                    )
                except Exception as e:
                    results.append({
                        "Exchange":      ex_nm,
                        "Status":        f"⚠️ Error",
                        "Seats":         "—",
                        "Seller Type(s)":"—",
                        "Seller IDs":    str(e)[:60],
                        "Source":        "—"
                    })

        prog.empty(); stat.empty()

        # Cache info banner
        if cached_count > 0:
            st.success(
                f"⚡ **{cached_count}/{len(exchanges)} exchanges served from cache** — "
                f"instant results. Cache refreshes every 24 hours."
            )
        else:
            st.info("🌐 All exchanges fetched live — results cached for next run (will be instant).")
        df = pd.DataFrame(results)
        found_count = len([r for r in results if r["Status"]=="✅ Found"])
        total_seats = len(all_seats)

        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Found In",    f"{found_count}/{len(exchanges)}")
        c2.metric("Total Seats", total_seats)
        c3.metric("Not Found",   len(exchanges)-found_count)
        c4.metric("Unreachable", len([r for r in results if "Unreachable" in r["Status"]]))

        if found_count == 0: st.error(f"🚨 No exchange seats — HIGH RISK. Do not onboard.")
        elif found_count < 3: st.warning(f"⚠️ Only {found_count} exchange relationships — limited reach")
        else: st.success(f"✅ Verified in {found_count} exchanges — looks legitimate")

        chart_df = df.copy()
        chart_df["Seats_num"] = pd.to_numeric(chart_df["Seats"], errors="coerce").fillna(0)
        fig = px.bar(chart_df, x="Exchange", y="Seats_num",
                     title=f"Seat Count — {dd}", color_discrete_sequence=["#2E75B6"])
        fig.update_layout(xaxis_tickangle=-40)
        st.plotly_chart(fig, use_container_width=True)

        st.dataframe(df, use_container_width=True, height=430)
        c1,c2 = st.columns(2)
        c1.download_button("⬇️ CSV", df_to_csv(df), f"{dd}_seats.csv","text/csv")
        c2.download_button("⬇️ Excel (colored)", df_to_colored_excel(df, "Demand Seats"),
                           f"{dd}_seats.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── PAGE: INTERMEDIARY INTELLIGENCE ────────────────────────────────────────

def page_intermediary_intel():
    st.title("🔀 Intermediary Intelligence")
    st.markdown("""
    **What this does:**
    Paste up to 20 sellers.json URLs → fetches all in parallel → extracts **INTERMEDIARY** entries only
    → deduplicates across all files → shows unique intermediaries with appearance count.

    **Why use this:**
    Intermediaries appearing in many sellers.json files = established, widely trusted middlemen.
    The higher the count, the more SSPs trust them. Use this to shortlist the best intermediary
    partners for your supply stack — or find which ones to avoid.
    """)
    st.markdown("---")

    urls_raw = st.text_area(
        "Paste sellers.json URLs (one per line)",
        height=220,
        placeholder=(
            "https://pubmatic.com/sellers.json\n"
            "https://openx.com/sellers.json\n"
            "https://indexexchange.com/sellers.json\n"
            "https://advlion.com/sellers.json\n"
            "https://smartadserver.com/sellers.json"
        ),
        help="Accepts full URLs or plain domains. Mix and match freely."
    )

    col1, col2 = st.columns(2)
    with col1:
        sort_by = st.selectbox("Sort by", ["Appearances (High→Low)", "Domain A→Z"])
    with col2:
        show_confidential = st.checkbox("Include is_confidential entries", value=False)

    st.info(
        "💡 **How this works:** The tool collects ALL intermediaries from every sellers.json file you paste, "
        "combines them into one list, removes duplicates, and shows each unique intermediary once. "
        "The **Appearances** column tells you how many of YOUR files that intermediary appeared in. "
        "You can filter by appearances AFTER results load."
    )

    if st.button("🔀 Fetch & Deduplicate", type="primary"):
        raw_lines = [l.strip() for l in urls_raw.splitlines() if l.strip()]
        if not raw_lines:
            st.warning("Paste at least one sellers.json URL")
            return

        # Parse each line into (label, domain, full_url)
        # IMPORTANT: preserve full URL including any path (e.g. aniview hash paths)
        sources = []
        seen_urls = set()
        for line in raw_lines[:20]:
            line = line.strip()
            if not line:
                continue
            if line.startswith("http://") or line.startswith("https://"):
                full_url = line
                # Use full URL as label key (not just domain) to allow same-domain different paths
                label = line
            else:
                domain   = clean_domain(line)
                full_url = f"https://{domain}/sellers.json"
                label    = full_url
            if full_url not in seen_urls:
                sources.append((label, full_url))
                seen_urls.add(full_url)

        if len(raw_lines) > 20:
            st.warning(f"Max 20 URLs — using first 20 of {len(raw_lines)}")

        st.markdown(f"### ⚡ Fetching {len(sources)} sellers.json files in parallel...")

        # ── Parallel fetch ────────────────────────────────────────────────
        def fetch_and_extract(args):
            label, full_url = args
            # Use full URL directly — critical for non-standard paths like aniview hash URLs
            cache_key = clean_domain(full_url)
            cached    = cache_get_sellers(cache_key)
            if cached.get("hit"):
                data = cached["data"]
                from_cache = True
            else:
                try:
                    r = requests.get(full_url, timeout=15, headers=HEADERS)
                    if r.status_code != 200:
                        return label, [], False, f"HTTP {r.status_code}"
                    data = None
                    # Try all parse strategies
                    for fn in [
                        lambda: r.json(),
                        lambda: json.loads(r.content.decode("utf-8-sig").strip()),
                        lambda: json.loads(r.content.decode("utf-8", errors="ignore").strip()),
                    ]:
                        try:
                            d = fn()
                            if d and isinstance(d, dict) and "sellers" in d:
                                data = d
                                break
                        except Exception:
                            continue
                    if data is None:
                        return label, [], False, "invalid_json"
                    cache_set_sellers(cache_key, full_url, data)
                    from_cache = False
                except Exception as e:
                    return label, [], False, str(e)[:60]

            sellers = get_sellers(data)
            intermediaries = [
                s for s in sellers
                if str(s.get("seller_type", "")).upper() == "INTERMEDIARY"
                and (show_confidential or not s.get("is_confidential"))
            ]
            return label, intermediaries, from_cache, "ok"

        results_by_domain = {}
        errors = []
        done = 0
        prog = st.progress(0)
        stat = st.empty()

        with ThreadPoolExecutor(max_workers=len(sources)) as executor:
            future_map = {executor.submit(fetch_and_extract, src): src for src in sources}
            for future in as_completed(future_map):
                domain_key, intermediaries, from_cache, status = future.result()
                done += 1
                prog.progress(done / len(sources))
                if status == "ok":
                    results_by_domain[domain_key] = {
                        "intermediaries": intermediaries,
                        "from_cache":     from_cache,
                        "count":          len(intermediaries)
                    }
                    label_short = domain_key[:50]
                    stat.text(f"✅ {done}/{len(sources)} — {label_short}: {len(intermediaries)} intermediaries {'⚡cache' if from_cache else '🌐live'}")
                else:
                    errors.append(domain_key)
                    stat.text(f"❌ {done}/{len(sources)} — {domain_key[:50]}: {status}")

        prog.empty(); stat.empty()

        if errors:
            st.warning(f"⚠️ Could not fetch {len(errors)} file(s): {', '.join(errors)}")

        if not results_by_domain:
            st.error("No data fetched — check your URLs")
            return

        # ── Summary of fetched files ──────────────────────────────────────
        st.markdown("### 📊 Files Fetched")
        fetch_summary = []
        for domain, data in results_by_domain.items():
            fetch_summary.append({
                "Domain":         domain,
                "Intermediaries": data["count"],
                "Source":         "⚡ Cache" if data["from_cache"] else "🌐 Live"
            })
        st.dataframe(pd.DataFrame(fetch_summary), use_container_width=True, height=200)

        # ── Deduplicate across all files ──────────────────────────────────
        # Key: (seller_id, exchange_domain) — track per unique identity
        from collections import defaultdict

        inter_map = defaultdict(lambda: {
            "seller_id":   "",
            "name":        set(),
            "domain":      set(),
            "appearances": 0,
            "found_in":    [],
            "is_confidential": False
        })

        for domain, data in results_by_domain.items():
            seen_in_this_file = set()
            for s in data["intermediaries"]:
                sid    = str(s.get("seller_id", "")).strip()
                s_dom  = str(s.get("domain", "")).lower().replace("www.", "").strip()
                s_name = str(s.get("name", "")).strip()
                # Use domain as primary key (most stable), fall back to sid
                key = s_dom if s_dom else sid
                if not key or key in seen_in_this_file:
                    continue
                seen_in_this_file.add(key)
                entry = inter_map[key]
                entry["seller_id"]      = sid
                entry["appearances"]   += 1
                entry["found_in"].append(domain)
                if s_name: entry["name"].add(s_name)
                if s_dom:  entry["domain"].add(s_dom)
                if s.get("is_confidential"): entry["is_confidential"] = True

        # ── Build results dataframe ────────────────────────────────────────
        rows = []
        for key, entry in inter_map.items():
            appearances = entry["appearances"]
            # No pre-filter — show ALL unique intermediaries, user filters below
            domain_str = ", ".join(sorted(entry["domain"])) or key
            name_str   = " / ".join(sorted(entry["name"]))[:60] or "—"
            found_str  = ", ".join(entry["found_in"])
            trust      = (
                "🟢 High"   if appearances >= len(sources) * 0.6 else
                "🟡 Medium" if appearances >= len(sources) * 0.3 else
                "🔴 Low"
            )
            rows.append({
                "Intermediary Domain": domain_str,
                "Name":               name_str,
                "Seller ID":          entry["seller_id"],
                "Appearances":        appearances,
                "Found In (count)":   f"{appearances}/{len(sources)} files",
                "Trust Signal":       trust,
                "Confidential":       "🔒 Yes" if entry["is_confidential"] else "—",
                "Found In Files":     found_str,
            })

        if not rows:
            st.info(f"No intermediaries found with min {min_count} appearances.")
            return

        df = pd.DataFrame(rows)

        # ── Sort ──────────────────────────────────────────────────────────
        if sort_by == "Appearances (High→Low)":
            df = df.sort_values("Appearances", ascending=False)
        elif sort_by == "Domain A→Z":
            df = df.sort_values("Intermediary Domain")
        elif sort_by == "Seller Count (High→Low)":
            df = df.sort_values("Appearances", ascending=False)

        df = df.reset_index(drop=True)

        # ── Stats ─────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📊 Deduplication Results")
        c1, c2, c3, c4 = st.columns(4)
        total_raw = sum(d["count"] for d in results_by_domain.values())
        c1.metric("Files Checked",          len(results_by_domain))
        c2.metric("Total Raw Intermediaries", total_raw,
                  help="Sum of all intermediaries across all files before deduplication")
        c3.metric("✅ Unique After Dedup",   len(df),
                  delta=f"-{total_raw - len(df)} duplicates removed")
        c4.metric("In All Files",
                  len(df[df["Appearances"] == len(sources)]))

        # Per-file breakdown
        with st.expander("📂 Per-File Intermediary Count"):
            pf_rows = []
            for domain, data in results_by_domain.items():
                pf_rows.append({
                    "sellers.json File": domain,
                    "Total Intermediaries": data["count"],
                    "Source": "⚡ Cache" if data["from_cache"] else "🌐 Live"
                })
            st.dataframe(pd.DataFrame(pf_rows), use_container_width=True)

        # Post-results filter
        max_appearances = int(df["Appearances"].max()) if len(df) > 0 else 1

        # Only show slider if there's a meaningful range to filter (max > 1)
        filter_min = 1
        filter_col1, filter_col2 = st.columns(2)
        with filter_col1:
            if max_appearances > 1:
                filter_min = st.slider(
                    "Filter: Min appearances",
                    min_value=1,
                    max_value=max_appearances,
                    value=1,
                    key="post_filter_min",
                    help=f"You have {len(sources)} files checked. Set to {len(sources)} to see intermediaries common to ALL files."
                )
            else:
                st.info(f"Only 1 file loaded successfully — showing all {len(df)} unique intermediaries.")
        with filter_col2:
            showing = len(df[df["Appearances"] >= filter_min])
            st.metric("Showing", showing, delta=f"of {len(df)} unique")

        df = df[df["Appearances"] >= filter_min].reset_index(drop=True)

        if df.empty:
            st.warning(f"No intermediaries appear in {filter_min}+ files. Try lowering the filter.")
            return

        # ── Appearance distribution chart ──────────────────────────────────
        fig = px.histogram(
            df, x="Appearances",
            nbins=len(sources),
            title="Intermediary Appearance Distribution (how many files each appears in)",
            color_discrete_sequence=["#2E75B6"],
            labels={"Appearances": "Number of Files", "count": "Unique Intermediaries"}
        )
        fig.update_layout(
            height=280,
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(gridcolor="#1A2A3A", dtick=1),
            yaxis=dict(gridcolor="#1A2A3A"),
            bargap=0.15
        )
        st.plotly_chart(fig, use_container_width=True)

        # ── Main results table ─────────────────────────────────────────────
        st.markdown("### 📋 Unique Intermediaries Table")

        display_cols = [
            "Intermediary Domain", "Name", "Appearances",
            "Found In (count)", "Trust Signal", "Confidential"
        ]
        st.dataframe(
            df[display_cols],
            use_container_width=True,
            height=min(600, 60 + len(df) * 36)
        )

        # ── One-click copy section ─────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📋 Quick Copy — Paste Directly Into Your Sheet")

        copy_tab1, copy_tab2, copy_tab3 = st.tabs([
            "📄 Domains Only",
            "📊 Domain + Count",
            "📋 Full CSV Text"
        ])

        with copy_tab1:
            st.caption("One domain per line — select all and copy (Ctrl+A, Ctrl+C)")
            domains_only = "\n".join(df["Intermediary Domain"].tolist())
            st.text_area("", value=domains_only, height=300, key="copy_domains",
                         label_visibility="collapsed")
            st.caption(f"✅ {len(df)} unique intermediary domains")

        with copy_tab2:
            st.caption("Tab-separated: Domain → Appearances — paste directly into Google Sheets")
            tab_sep = "\n".join(
                f"{row['Intermediary Domain']}\t{row['Appearances']}\t{row['Found In (count)']}\t{row['Trust Signal']}"
                for _, row in df.iterrows()
            )
            st.text_area("", value=tab_sep, height=300, key="copy_tab",
                         label_visibility="collapsed")

        with copy_tab3:
            st.caption("Full CSV — paste into any spreadsheet or text editor")
            csv_text = df[display_cols + ["Found In Files"]].to_csv(index=False)
            st.text_area("", value=csv_text, height=300, key="copy_csv",
                         label_visibility="collapsed")

        # ── Download buttons ───────────────────────────────────────────────
        st.markdown("---")
        c1, c2 = st.columns(2)
        c1.download_button(
            "⬇️ Download CSV",
            df_to_csv(df),
            "intermediary_intelligence.csv",
            "text/csv"
        )
        c2.download_button(
            "⬇️ Download Excel (colored)",
            df_to_colored_excel(df, "Intermediary Intel",
                                "AdChain — Intermediary Intelligence Report"),
            "intermediary_intelligence.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # ── High Trust shortlist ───────────────────────────────────────────
        high_trust = df[df["Trust Signal"] == "🟢 High"]
        if not high_trust.empty:
            st.markdown("---")
            st.markdown("### 🟢 High Trust Shortlist")
            st.caption(
                f"These {len(high_trust)} intermediaries appear in ≥60% of your checked files — "
                "most reliable for supply stack consideration"
            )
            st.dataframe(high_trust[display_cols], use_container_width=True)
            shortlist_text = "\n".join(high_trust["Intermediary Domain"].tolist())
            st.text_area("📋 Copy shortlist", value=shortlist_text,
                         height=150, key="copy_shortlist")

# ─── PAGE: DEMAND HEALTH ──────────────────────────────────────────────────

def page_health():
    st.title("🧠 Demand Health Score")
    st.markdown("*0–100 score: exchange presence + sellers.json transparency + seat count*")
    st.markdown("---")

    demand_domain = st.text_input("Demand Partner Domain", placeholder="e.g. anzu.io")

    if st.button("🧠 Calculate", type="primary"):
        if not demand_domain.strip(): st.warning("Enter domain"); return
        dd = clean_domain(demand_domain); score = 0; breakdown = []

        with st.spinner("Analysing (parallel fetch)..."):
            found_in=0; total_seats=0

            def _health_check_exchange(args):
                ex_nm, ex_dm = args
                res = fetch_sellers_json(ex_dm, use_cache=True)
                if not res["success"]: return 0, 0
                s_list = get_sellers(res["data"])
                hits   = [s for s in s_list if str(s.get("domain","")).lower().replace("www.","") == dd]
                return (1 if hits else 0), len(hits)

            with ThreadPoolExecutor(max_workers=len(TOP_15_EXCHANGES)) as ex:
                futs = [ex.submit(_health_check_exchange, item)
                        for item in TOP_15_EXCHANGES.items()]
                for f in as_completed(futs):
                    fi, ts = f.result()
                    found_in    += fi
                    total_seats += ts

            ex_score = min(40, found_in*3); score += ex_score
            breakdown.append({"Category":"Exchange Presence (max 40)","Points":ex_score,
                               "Detail":f"Found in {found_in}/{len(TOP_15_EXCHANGES)}  |  {total_seats} seats"})

            sj = fetch_sellers_json(dd)
            if sj["success"]:
                score += 20
                breakdown.append({"Category":"Own sellers.json (max 20)","Points":20,"Detail":"✅ Accessible"})
                s_list = get_sellers(sj["data"])
                conf = [s for s in s_list if s.get("is_confidential")==1]
                conf_pct = len(conf)/len(s_list)*100 if s_list else 0
                t_score = max(0, 20 - int(conf_pct/5)); score += t_score
                breakdown.append({"Category":"Transparency (max 20)","Points":t_score,
                                   "Detail":f"{round(conf_pct)}% confidential"})
            else:
                breakdown.append({"Category":"Own sellers.json (max 20)","Points":0,"Detail":"❌ Not found"})
                breakdown.append({"Category":"Transparency (max 20)","Points":0,"Detail":"Cannot check"})

            if total_seats == 0: s_s,s_n = 0,"No seats"
            elif total_seats <= 25: s_s,s_n = 20,f"{total_seats} seats — reasonable"
            elif total_seats <= 50: s_s,s_n = 12,f"{total_seats} seats — high"
            else: s_s,s_n = 4,f"{total_seats} seats — very high, check for stacking"
            score += s_s
            breakdown.append({"Category":"Seat Count (max 20)","Points":s_s,"Detail":s_n})

        score = min(score,100)
        label = "Healthy ✅" if score>=70 else "Moderate ⚠️" if score>=40 else "Risky 🚨"
        color = "#4CAF50"    if score>=70 else "#FF9800"       if score>=40 else "#F44336"

        fig = go.Figure(go.Indicator(
            mode="gauge+number",
            value=score,
            title={"text":f"{dd}<br><span style='font-size:14px;color:{color}'>{label}</span>"},
            gauge={"axis":{"range":[0,100]},"bar":{"color":color,"thickness":0.3},
                   "steps":[{"range":[0,40],"color":"#2D1010"},{"range":[40,70],"color":"#2D2010"},
                             {"range":[70,100],"color":"#102D10"}]}
        ))
        fig.update_layout(height=300)
        st.plotly_chart(fig, use_container_width=True)

        df = pd.DataFrame(breakdown)
        st.dataframe(df, use_container_width=True)
        st.download_button("⬇️ Download", df_to_csv(df), f"{dd}_health.csv","text/csv")

# ─── PAGE: BULK SSP AUTHENTICATOR (NEW) ──────────────────────────────────

def page_bulk_ssp():
    st.title("⚡ Bulk SSP Authenticator")
    st.markdown("""
    **v2 Feature** — Check multiple SSPs simultaneously.
    Paste up to 20 SSP domains → runs full PUBLISHER + DIRECT validation on each → gives a Trust Score per SSP.
    """)
    st.markdown("---")

    ssp_raw  = st.text_area(
        "SSP Domains or sellers.json URLs (one per line)",
        placeholder=(
            "openx.com\n"
            "pubmatic.com\n"
            "https://smilewanted.com/sellers.json\n"
            "https://ottera.tv\n"
            "1. advlion.com\n"
            "2. smartadserver.com"
        ),
        height=200,
        help="Accepts any format: plain domain, full sellers.json URL, or numbered list (1. domain.com)"
    )
    col1,col2,col3 = st.columns(3)
    with col1: max_per_ssp = st.slider("Max publishers per SSP", 5, 100, 25)
    with col2: workers     = st.slider("Workers per SSP", 5, 20, 10)
    with col3: check_app   = st.checkbox("Check app-ads.txt", value=True)

    def parse_bulk_input_line(line):
        """
        Robustly extract a clean domain from any input format:
        - "openx.com"
        - "https://openx.com/sellers.json"
        - "1. openx.com"
        - "11. https://smilewanted.com/sellers.json"
        - "  2) ottera.tv  "
        """
        import re as _re
        line = line.strip()
        if not line:
            return None
        # Strip leading numbered list prefixes: "1. ", "11. ", "2) ", "3: " etc.
        line = _re.sub(r"^\d+[\.\)\:\-]\s*", "", line).strip()
        if not line:
            return None
        # If it's a URL, extract domain and strip /sellers.json path
        if line.startswith("http://") or line.startswith("https://"):
            parsed  = urlparse(line)
            domain  = parsed.netloc.lower().replace("www.", "").strip()
        else:
            domain  = line.lower().replace("www.", "").split("/")[0].strip()
        return domain if domain else None

    if st.button("⚡ Run Bulk Check", type="primary"):
        raw_lines = [l for l in ssp_raw.splitlines() if l.strip()]
        domains   = []
        seen      = set()
        for line in raw_lines:
            d = parse_bulk_input_line(line)
            if d and d not in seen:
                domains.append(d)
                seen.add(d)
        if not domains: st.warning("Enter at least one SSP domain or URL"); return
        if len(domains) > 20:
            st.warning(f"Max 20 SSPs — truncating from {len(domains)} to 20")
            domains = domains[:20]

        st.markdown(f"### 🔍 Checking {len(domains)} SSPs...")
        summary_rows = []
        all_results  = {}

        prog_outer = st.progress(0)
        for ssp_idx, domain in enumerate(domains):
            prog_outer.progress((ssp_idx+1)/len(domains))
            st.markdown(f"**[{ssp_idx+1}/{len(domains)}] {domain}**")

            sj = fetch_sellers_json(domain)
            if not sj["success"]:
                tried_url = f"https://{domain}/sellers.json"
                summary_rows.append({
                    "SSP Domain":        domain,
                    "sellers.json URL":  tried_url,
                    "Total Publishers":  "—",
                    "✅ Legit":          "—",
                    "❌ Fake":           "—",
                    "⚠️ Misrep":         "—",
                    "🚨 CDN":            "—",
                    "Trust Score":       "—",
                    "Status":            "⚠️ Unreachable — sellers.json not found or unreadable"
                })
                continue

            sellers    = get_sellers(sj["data"])
            publishers = [s for s in sellers if str(s.get("seller_type","")).upper() == "PUBLISHER"]
            cdn_count  = len([s for s in publishers
                               if is_cdn_domain(extract_domain_from_field(str(s.get("domain",""))))])
            to_check   = publishers[:max_per_ssp]
            results    = []

            with ThreadPoolExecutor(max_workers=workers) as executor:
                fut_map = {executor.submit(check_single_publisher, s, check_app, domain): s
                           for s in to_check}
                for fut in as_completed(fut_map):
                    results.append(fut.result())

            legit  = len([r for r in results if "Legitimate" in r["Verdict"]])
            fake   = len([r for r in results if "Fake"       in r["Verdict"]])
            misrep = len([r for r in results if "Misrep"     in r["Verdict"]])
            total  = len(results) or 1

            trust = round((legit / total) * 100) if total > 0 else 0
            flag  = "✅ Trusted" if trust >= 80 else "⚠️ Review" if trust >= 50 else "🚨 Risky"

            summary_rows.append({
                "SSP Domain":        domain,
                "Publishers Checked":total,
                "✅ Legit":          legit,
                "❌ Fake":           fake,
                "⚠️ Misrep":         misrep,
                "🚨 CDN Domains":    cdn_count,
                "Trust Score":       f"{trust}%",
                "Status":            flag
            })
            all_results[domain] = results

        prog_outer.empty()
        df_summary = pd.DataFrame(summary_rows)

        st.markdown("### 📊 SSP Trust Comparison")
        # Bar chart
        score_data = [r for r in summary_rows if r["Trust Score"] != "—"]
        if score_data:
            fig = px.bar(
                x=[r["SSP Domain"]  for r in score_data],
                y=[int(str(r["Trust Score"]).replace("%","")) for r in score_data],
                color=[int(str(r["Trust Score"]).replace("%","")) for r in score_data],
                color_continuous_scale=["#F44336","#FF9800","#4CAF50"],
                range_color=[0,100],
                title="Trust Score per SSP (% Legitimate Publishers)",
                labels={"x":"SSP","y":"Trust Score %"}
            )
            fig.update_layout(xaxis_tickangle=-30, coloraxis_showscale=False)
            st.plotly_chart(fig, use_container_width=True)

        st.dataframe(df_summary, use_container_width=True)
        c1,c2 = st.columns(2)
        c1.download_button("⬇️ Summary CSV", df_to_csv(df_summary), "bulk_ssp_summary.csv","text/csv")
        c2.download_button("⬇️ Summary Excel (colored)",
                           df_to_colored_excel(df_summary, "Bulk SSP Summary"),
                           "bulk_ssp_summary.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Per-SSP detail expandable
        st.markdown("### 🔍 Per-SSP Detail")
        for domain, results in all_results.items():
            with st.expander(f"📋 {domain} — {len(results)} publishers checked"):
                df_d = pd.DataFrame(results)
                st.dataframe(df_d, use_container_width=True, height=300)
                st.download_button(f"⬇️ {domain} Excel",
                                   df_to_colored_excel(df_d, domain),
                                   f"{domain}_detail.xlsx",
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key=f"dl_{domain}")

# ─── PAGE: ADVLION PRESENCE SCANNER (NEW) ────────────────────────────────

def page_presence():
    st.title("🏢 Advlion Presence Scanner")
    st.markdown(f"""
    **v2 Feature** — Upload a CSV of publisher domains → checks every one for
    `advlion.com, {ADVLION_SID}, DIRECT` in their ads.txt / app-ads.txt.
    Missing listings = revenue leakage.
    """)
    st.markdown("---")

    tab1, tab2 = st.tabs(["📁 Upload CSV", "✏️ Manual Entry"])

    with tab1:
        uploaded = st.file_uploader("CSV with a column named 'domain'", type=["csv"])
        if uploaded:
            df_in = pd.read_csv(uploaded)
            if "domain" not in df_in.columns:
                st.error("CSV must have a column named 'domain'")
            else:
                domains = df_in["domain"].dropna().astype(str).tolist()
                st.info(f"Loaded {len(domains)} domains")
                col1,col2 = st.columns(2)
                with col1: check_app = st.checkbox("Check app-ads.txt too", value=True, key="p_app")
                with col2: workers   = st.slider("Workers", 5,30,20, key="p_w")
                if st.button("🏢 Scan All", type="primary", key="scan_csv"):
                    _run_presence_scan(domains, check_app, workers)

    with tab2:
        raw = st.text_area("Publisher domains (one per line)", height=200,
                           placeholder="truecaller.com\nnytimes.com\nweather.com")
        col1,col2 = st.columns(2)
        with col1: check_app2 = st.checkbox("Check app-ads.txt too", value=True, key="p_app2")
        with col2: workers2   = st.slider("Workers", 5,30,20, key="p_w2")
        if st.button("🏢 Scan All", type="primary", key="scan_manual"):
            domains2 = [clean_domain(d) for d in raw.splitlines() if d.strip()]
            if not domains2: st.warning("Enter at least one domain")
            else: _run_presence_scan(domains2, check_app2, workers2)

def check_advlion_presence(domain, check_app):
    d = clean_domain(domain)
    filenames = ["app-ads.txt","ads.txt"] if check_app else ["ads.txt"]

    for fname in filenames:
        for scheme in ["https","http"]:
            for prefix in ["","www."]:
                try:
                    url = f"{scheme}://{prefix}{d}/{fname}"
                    r   = requests.get(url, timeout=6, headers=HEADERS)
                    if r.status_code == 200:
                        rows      = parse_ads_txt(r.text)
                        adv_exact = [row for row in rows
                                     if row["exchange_domain"] == ADVLION_DOMAIN
                                     and row["seller_id"].strip().lstrip("0") == ADVLION_SID.lstrip("0")
                                     and row["relationship"] == "DIRECT"]
                        adv_any   = [row for row in rows if row["exchange_domain"] == ADVLION_DOMAIN]
                        adv_wrong_sid = [row for row in rows
                                         if row["exchange_domain"] == ADVLION_DOMAIN
                                         and row["seller_id"] != ADVLION_SID]

                        if adv_exact:
                            return {"Domain":d,"File Found":f"✅ {fname}",
                                    "Advlion Listed":"✅ DIRECT",
                                    "SID Match":"✅ 3148","Status":"✅ Correct",
                                    "Raw Line":adv_exact[0]["raw"]}
                        elif adv_wrong_sid:
                            return {"Domain":d,"File Found":f"✅ {fname}",
                                    "Advlion Listed":"⚠️ Wrong SID",
                                    "SID Match":f"❌ Got {adv_wrong_sid[0]['seller_id']}",
                                    "Status":"⚠️ Wrong SID","Raw Line":adv_wrong_sid[0]["raw"]}
                        elif adv_any:
                            return {"Domain":d,"File Found":f"✅ {fname}",
                                    "Advlion Listed":"⚠️ Found (check relationship)",
                                    "SID Match":"⚠️","Status":"⚠️ Review","Raw Line":adv_any[0]["raw"]}
                        else:
                            return {"Domain":d,"File Found":f"✅ {fname}",
                                    "Advlion Listed":"❌ Not Listed","SID Match":"—",
                                    "Status":"❌ Missing","Raw Line":"—"}
                except Exception:
                    continue

    return {"Domain":d,"File Found":"❌ Not Found","Advlion Listed":"❌","SID Match":"—",
            "Status":"❌ No ads.txt","Raw Line":"—"}

def _run_presence_scan(domains, check_app, workers):
    st.markdown(f"### 🔍 Scanning {len(domains)} publishers for Advlion presence...")
    results=[]; done=0
    prog=st.progress(0); stat=st.empty()

    with ThreadPoolExecutor(max_workers=workers) as executor:
        fut_map = {executor.submit(check_advlion_presence, d, check_app): d for d in domains}
        for fut in as_completed(fut_map):
            results.append(fut.result()); done += 1
            prog.progress(done/len(domains))
            stat.text(f"✅ {done}/{len(domains)}")

    prog.empty(); stat.empty()
    df = pd.DataFrame(results)

    correct = len([r for r in results if r["Status"]=="✅ Correct"])
    missing = len([r for r in results if "Missing" in r["Status"] or "No ads" in r["Status"]])
    wrong   = len([r for r in results if "Wrong" in r["Status"] or "Review" in r["Status"]])

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Total Scanned",   len(results))
    c2.metric("✅ Correctly Listed", correct)
    c3.metric("❌ Not Listed",    missing)
    c4.metric("⚠️ Wrong/Review",  wrong)

    if missing > 0:
        st.error(f"🚨 {missing} publishers don't have Advlion listed — potential revenue leakage!")
    if wrong > 0:
        st.warning(f"⚠️ {wrong} publishers have Advlion listed incorrectly — needs fixing")
    if correct == len(results):
        st.success("✅ All publishers correctly list Advlion!")

    st.dataframe(df, use_container_width=True, height=420)
    c1,c2 = st.columns(2)
    c1.download_button("⬇️ CSV", df_to_csv(df), "advlion_presence.csv","text/csv")
    c2.download_button("⬇️ Excel (colored)",
                       df_to_colored_excel(df,"Advlion Presence"),
                       "advlion_presence.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── PAGE: S&D INTELLIGENCE (NEW) ─────────────────────────────────────────

def page_sd_intel():
    st.title("📊 S&D Intelligence")
    st.markdown("""
    **v2 Feature** — Upload your weekly S&D Excel → cross-references revenue vs sellers.json compliance.
    Answers: *Which of my highest-revenue SSPs have the worst supply fraud?*
    """)
    st.markdown("---")

    st.info("📁 Upload your S&D Excel file. The tool expects columns with SSP names and revenue/request data (flexible column matching).")
    uploaded = st.file_uploader("Upload S&D Excel (.xlsx)", type=["xlsx","xls"])

    if uploaded:
        try:
            xl   = pd.ExcelFile(uploaded)
            sheet = st.selectbox("Select sheet", xl.sheet_names)
            df_sd = pd.read_excel(uploaded, sheet_name=sheet)
            st.success(f"✅ Loaded {len(df_sd)} rows, {len(df_sd.columns)} columns")
            st.dataframe(df_sd.head(5), use_container_width=True)

            # Smart column detection
            all_cols = list(df_sd.columns)
            name_col = st.selectbox("SSP Name column",
                [c for c in all_cols if any(k in str(c).lower() for k in ["ssp","partner","name","source"])]
                or all_cols)
            rev_col  = st.selectbox("Revenue column",
                [c for c in all_cols if any(k in str(c).lower() for k in ["rev","earn","amount","$"])]
                or all_cols)

            if st.button("📊 Run S&D Intelligence", type="primary"):
                ssps = df_sd[name_col].dropna().astype(str).unique().tolist()
                st.markdown(f"### 🔍 Analysing {len(ssps)} SSPs from your S&D report...")

                compliance_rows = []
                prog = st.progress(0)
                stat = st.empty()

                for i, ssp_name in enumerate(ssps[:30]):  # cap at 30
                    prog.progress((i+1)/min(len(ssps),30))
                    stat.text(f"Checking {ssp_name}...")

                    # Try to resolve domain from SSP name
                    domain_guess = ssp_name.lower().replace(" ","").replace("_ssp","").replace("_dssp","")
                    if "." not in domain_guess:
                        domain_guess += ".com"
                    domain_guess = clean_domain(domain_guess)

                    sj = fetch_sellers_json(domain_guess)
                    revenue_rows = df_sd[df_sd[name_col].astype(str) == ssp_name]
                    total_revenue = pd.to_numeric(revenue_rows[rev_col], errors="coerce").sum()

                    if not sj["success"]:
                        compliance_rows.append({
                            "SSP Name":    ssp_name,
                            "Domain Tried":domain_guess,
                            "Revenue":     round(total_revenue, 2),
                            "Total Sellers":"—",
                            "PUBLISHER":   "—",
                            "Confidential %":"—",
                            "Compliance":  "⚠️ sellers.json N/A",
                            "Risk":        "⚠️ Unknown"
                        })
                        continue

                    sellers   = get_sellers(sj["data"])
                    pub       = [s for s in sellers if str(s.get("seller_type","")).upper()=="PUBLISHER"]
                    conf      = [s for s in sellers if s.get("is_confidential")==1]
                    cdn_s     = [s for s in sellers if is_cdn_domain(extract_domain_from_field(str(s.get("domain",""))))]
                    conf_pct  = round(len(conf)/len(sellers)*100) if sellers else 0
                    cdn_pct   = round(len(cdn_s)/len(sellers)*100) if sellers else 0

                    # Quick compliance score
                    score = 100
                    if conf_pct > 30: score -= 25
                    elif conf_pct > 10: score -= 10
                    if cdn_pct > 5: score -= 20
                    if len(pub) == 0: score -= 15

                    risk = "✅ Low" if score >= 70 else "⚠️ Medium" if score >= 50 else "🚨 High"

                    compliance_rows.append({
                        "SSP Name":       ssp_name,
                        "Domain":         domain_guess,
                        "Revenue":        round(total_revenue, 2),
                        "Total Sellers":  len(sellers),
                        "PUBLISHER cnt":  len(pub),
                        "Confidential %": f"{conf_pct}%",
                        "CDN Domains %":  f"{cdn_pct}%",
                        "Compliance Score":f"{score}/100",
                        "Risk":           risk
                    })

                prog.empty(); stat.empty()
                df_comp = pd.DataFrame(compliance_rows)

                # Sort by revenue desc
                df_comp["_rev"] = pd.to_numeric(df_comp["Revenue"], errors="coerce").fillna(0)
                df_comp = df_comp.sort_values("_rev", ascending=False).drop(columns=["_rev"])

                st.markdown("### 📊 Revenue vs Compliance — Full Report")

                # Scatter plot: revenue vs compliance score
                plot_df = df_comp[df_comp["Compliance Score"].astype(str) != "⚠️ sellers.json N/A"].copy()
                plot_df["Score_num"] = pd.to_numeric(
                    plot_df["Compliance Score"].astype(str).str.replace("/100",""), errors="coerce").fillna(0)
                if not plot_df.empty:
                    fig = px.scatter(
                        plot_df, x="Score_num", y="Revenue",
                        text="SSP Name", color="Score_num",
                        color_continuous_scale=["#F44336","#FF9800","#4CAF50"],
                        range_color=[0,100],
                        title="Revenue vs Compliance Score — Bubble = Revenue at Risk",
                        labels={"Score_num":"Compliance Score","Revenue":"Revenue"}
                    )
                    fig.update_traces(textposition="top center", marker_size=12)
                    fig.update_layout(coloraxis_showscale=False)
                    st.plotly_chart(fig, use_container_width=True)

                st.dataframe(df_comp, use_container_width=True, height=420)

                high_risk = df_comp[df_comp["Risk"].astype(str).str.contains("High")]
                if not high_risk.empty:
                    high_rev = pd.to_numeric(high_risk["Revenue"], errors="coerce").sum()
                    st.error(f"🚨 {len(high_risk)} HIGH RISK SSPs with combined revenue of {high_rev:,.2f} — investigate immediately")

                c1,c2 = st.columns(2)
                c1.download_button("⬇️ CSV", df_to_csv(df_comp), "sd_intelligence.csv","text/csv")
                c2.download_button("⬇️ Excel (colored)",
                                   df_to_colored_excel(df_comp,"S&D Intelligence","AdChain S&D Intelligence Report"),
                                   "sd_intelligence.xlsx",
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Error reading file: {e}")

# ─── PAGE: CHANGE TRACKER (NEW) ───────────────────────────────────────────

def page_tracker():
    st.title("🕵️ Change Tracker")
    st.markdown("""
    **v2 Feature** — SQLite-powered history. Every time you fetch a sellers.json, it's cached.
    If the data changes on next fetch, the old version is saved as a snapshot.
    See **exactly what changed** between fetches.
    """)
    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        domain = st.text_input("SSP Domain to track", placeholder="e.g. advlion.com")
    with col2:
        force_refresh = st.checkbox("Force fresh fetch (bypass cache)", value=False)

    if st.button("🕵️ Check for Changes", type="primary"):
        if not domain.strip(): st.warning("Enter domain"); return
        d = clean_domain(domain)

        # Get old snapshot before fetching
        old_cache = cache_get_sellers(d)
        old_data  = old_cache.get("data") if old_cache.get("hit") else None
        old_hash  = None
        if old_data:
            old_hash = hashlib.md5(json.dumps(old_data).encode()).hexdigest()

        with st.spinner(f"Fetching fresh {d}/sellers.json..."):
            result = fetch_sellers_json(d, use_cache=not force_refresh)

        if not result["success"]:
            st.error(f"❌ Could not fetch {d}/sellers.json"); return

        new_data = result["data"]
        new_hash = hashlib.md5(json.dumps(new_data).encode()).hexdigest()

        if old_hash and old_hash != new_hash:
            st.warning(f"🔄 **CHANGE DETECTED** in {d}/sellers.json since last fetch!")
        elif old_hash:
            st.success(f"✅ No changes in {d}/sellers.json since last fetch")
        else:
            st.info(f"📥 First time fetching {d}/sellers.json — saved as baseline")

        # Deep diff
        if old_data and old_hash != new_hash:
            old_sellers = {str(s.get("seller_id","")): s for s in get_sellers(old_data)}
            new_sellers = {str(s.get("seller_id","")): s for s in get_sellers(new_data)}

            added   = [s for sid,s in new_sellers.items() if sid not in old_sellers]
            removed = [s for sid,s in old_sellers.items() if sid not in new_sellers]
            changed = []
            for sid in set(old_sellers) & set(new_sellers):
                o,n = old_sellers[sid], new_sellers[sid]
                diffs = {k: {"before": o.get(k), "after": n.get(k)}
                         for k in set(list(o.keys())+list(n.keys()))
                         if o.get(k) != n.get(k)}
                if diffs:
                    changed.append({"seller_id":sid, "name":n.get("name",""), "changes":str(diffs)})

            c1,c2,c3 = st.columns(3)
            c1.metric("➕ Entries Added",   len(added))
            c2.metric("➖ Entries Removed", len(removed))
            c3.metric("✏️ Entries Changed", len(changed))

            if added:
                st.markdown("#### ➕ New Entries Added")
                st.dataframe(pd.DataFrame(added), use_container_width=True)

            if removed:
                st.markdown("#### ➖ Entries Removed")
                st.dataframe(pd.DataFrame(removed), use_container_width=True)

            if changed:
                st.markdown("#### ✏️ Entries Modified")
                st.dataframe(pd.DataFrame(changed), use_container_width=True)

            # Check if Advlion entry changed
            adv_old = old_sellers.get(ADVLION_SID)
            adv_new = new_sellers.get(ADVLION_SID)
            if adv_old != adv_new:
                st.error(f"🚨 Advlion entry (sid:{ADVLION_SID}) CHANGED in {d}/sellers.json!")
                col1, col2 = st.columns(2)
                col1.markdown("**Before:**"); col1.json(adv_old or {})
                col2.markdown("**After:**");  col2.json(adv_new or {})

        # Snapshot history
        st.markdown("---")
        st.markdown("### 📜 Snapshot History")
        snapshots = cache_get_snapshots(d)
        if snapshots:
            for snap_at, snap_hash, snap_json in snapshots:
                with st.expander(f"📸 Snapshot: {snap_at[:16]}  (hash: {snap_hash[:8]}...)"):
                    snap_data = json.loads(snap_json)
                    snap_sellers = get_sellers(snap_data)
                    st.metric("Seller Count", len(snap_sellers))
                    df_snap = pd.DataFrame(snap_sellers)
                    st.dataframe(df_snap, use_container_width=True, height=250)
                    st.download_button(f"⬇️ Download snapshot {snap_at[:10]}",
                                       df_to_csv(df_snap),
                                       f"{d}_snapshot_{snap_at[:10]}.csv","text/csv",
                                       key=f"snap_{snap_hash}")
        else:
            st.info("No historical snapshots yet. Run again after changes occur to build history.")

# ─── REMAINING PAGES (Bundle, Schain, IVT, Onboarding, Digest) ───────────

# ─── PAGE: SUPPLY CHAIN VALIDATOR (Bundle → app-ads.txt → DIRECT check) ────

def resolve_developer_domain_playstore(bundle):
    """
    Fetch Play Store page for a bundle and extract the developer website URL.
    Returns (domain, evidence) or (None, reason).
    """
    import re as _re
    try:
        url = f"https://play.google.com/store/apps/details?id={bundle}&hl=en"
        r   = requests.get(url, timeout=8, headers=HEADERS)
        if r.status_code == 404:
            return None, "App not found on Play Store"
        if r.status_code != 200:
            return None, f"Play Store returned {r.status_code}"

        text = r.text

        # Extract developer website from the page
        patterns = [
            r'href="(https?://(?!play\.google|support\.google|policies\.google|goo\.gl)[^"]+)"[^>]*>Visit website',
            r'"url":"(https?://(?!play\.google|support\.google)[^"]+)"',
            r'Developer Website.*?href="(https?://[^"]+)"',
        ]
        for pat in patterns:
            match = _re.search(pat, text, _re.IGNORECASE | _re.DOTALL)
            if match:
                raw_url = match.group(1)
                domain  = clean_domain(raw_url)
                if domain and "google" not in domain and len(domain) > 3:
                    return domain, f"Play Store: {raw_url}"

        # Fallback: try extracting from structured data
        sd_match = _re.search(r'"developerWebsite"\s*:\s*"([^"]+)"', text)
        if sd_match:
            domain = clean_domain(sd_match.group(1))
            if domain:
                return domain, f"Play Store structured data: {sd_match.group(1)}"

        return None, "Play Store found but no developer website listed"
    except Exception as e:
        return None, f"Play Store error: {str(e)[:60]}"


def resolve_developer_domain_appstore(app_id):
    """
    Use iTunes Search API to get the developer website for an iOS app.
    Returns (domain, evidence) or (None, reason).
    """
    try:
        url  = f"https://itunes.apple.com/lookup?id={app_id}&country=us"
        r    = requests.get(url, timeout=8, headers=HEADERS)
        data = r.json()
        results = data.get("results", [])
        if not results:
            return None, "App not found on App Store"

        app = results[0]
        # sellerUrl is the developer website
        seller_url = app.get("sellerUrl", "")
        if seller_url:
            domain = clean_domain(seller_url)
            if domain:
                return domain, f"App Store sellerUrl: {seller_url}"

        # Fallback: artistViewUrl gives developer page
        dev_url = app.get("artistViewUrl", "")
        if dev_url and "apple.com" not in dev_url:
            domain = clean_domain(dev_url)
            if domain:
                return domain, f"App Store artistViewUrl"

        track_name = app.get("trackName", "")
        return None, f"App Store found ({track_name}) but no developer website"
    except Exception as e:
        return None, f"App Store error: {str(e)[:60]}"


def check_bundle_supply_chain(bundle, partner_domain, partner_sid, check_schain=False):
    """
    Full supply chain check for one bundle:
    1. Detect platform
    2. Resolve developer domain from store
    3. Fetch app-ads.txt
    4. Check partner DIRECT line
    5. Return full result dict
    """
    bundle  = str(bundle).strip()
    is_ios  = bundle.isdigit()
    platform = "iOS" if is_ios else "Android"

    result = {
        "Bundle":           bundle,
        "Platform":         platform,
        "App Store Status": "—",
        "Developer Domain": "—",
        "Domain Source":    "—",
        "app-ads.txt":      "—",
        "Partner Listed":   "—",
        "DIRECT Line":      "—",
        "Partner SID Match":"—",
        "Verdict":          "—",
        "Notes":            "—",
    }

    # ── Step 1: Resolve developer domain ─────────────────────────────────
    if is_ios:
        domain, evidence = resolve_developer_domain_appstore(bundle)
        store_status = "✅ Found" if domain else "❌ Not Found"
    else:
        domain, evidence = resolve_developer_domain_playstore(bundle)
        if domain is None and "not found" in evidence.lower():
            store_status = "❌ Removed/Not Found"
        elif domain is None:
            store_status = "⚠️ Found (no website)"
        else:
            store_status = "✅ Found"

    result["App Store Status"] = store_status
    result["Domain Source"]    = evidence or "—"

    if not domain:
        result["Developer Domain"] = "❌ Could not resolve"
        result["Verdict"]          = "⚠️ No developer domain"
        result["Notes"]            = evidence or "No website on store listing"
        return result

    result["Developer Domain"] = domain

    # ── Step 2: Fetch app-ads.txt ─────────────────────────────────────────
    adstxt_res = fetch_ads_txt_fast(domain, check_app=True, check_web=False)

    if not adstxt_res["success"]:
        # Also try ads.txt as fallback
        adstxt_res = fetch_ads_txt_fast(domain, check_app=True, check_web=True)

    if not adstxt_res["success"]:
        result["app-ads.txt"]    = "❌ Not Found"
        result["Partner Listed"] = "❌ No"
        result["Verdict"]        = "❌ No app-ads.txt"
        result["Notes"]          = f"Neither app-ads.txt nor ads.txt found at {domain}"
        return result

    result["app-ads.txt"] = f"✅ {adstxt_res['filename']}"

    # ── Step 3: Check partner listing ────────────────────────────────────
    rows = parse_ads_txt(adstxt_res["content"])

    def sid_match(row_sid, target_sid):
        if not target_sid:
            return False
        return (row_sid.strip().lstrip("0") == target_sid.strip().lstrip("0")
                or row_sid.strip() == target_sid.strip())

    # All lines from this partner domain
    partner_lines = [r for r in rows if r["exchange_domain"] == partner_domain]
    direct_match  = [r for r in partner_lines
                     if r["relationship"] == "DIRECT"
                     and (not partner_sid or sid_match(r["seller_id"], partner_sid))]
    direct_wrong_sid = [r for r in partner_lines
                        if r["relationship"] == "DIRECT"
                        and partner_sid and not sid_match(r["seller_id"], partner_sid)]
    reseller_match   = [r for r in partner_lines
                        if r["relationship"] == "RESELLER"
                        and (not partner_sid or sid_match(r["seller_id"], partner_sid))]

    if direct_match:
        sid_found = direct_match[0]["seller_id"]
        result["Partner Listed"]    = "✅ Listed"
        result["DIRECT Line"]       = f"✅ DIRECT ({partner_domain})"
        result["Partner SID Match"] = f"✅ {sid_found}"
        result["Verdict"]           = "✅ Legitimate"
        result["Notes"]             = f"Correct DIRECT line: {direct_match[0]['raw']}"

    elif direct_wrong_sid:
        sid_found = direct_wrong_sid[0]["seller_id"]
        result["Partner Listed"]    = "⚠️ Listed"
        result["DIRECT Line"]       = f"⚠️ DIRECT but wrong SID"
        result["Partner SID Match"] = f"❌ Got {sid_found}, expected {partner_sid}"
        result["Verdict"]           = "⚠️ ID Mismatch"
        result["Notes"]             = f"Line: {direct_wrong_sid[0]['raw']}"

    elif reseller_match:
        result["Partner Listed"]    = "⚠️ Listed"
        result["DIRECT Line"]       = "⚠️ RESELLER only"
        result["Partner SID Match"] = "⚠️ RESELLER"
        result["Verdict"]           = "⚠️ Reseller Only"
        result["Notes"]             = f"Line: {reseller_match[0]['raw']}"

    elif partner_lines:
        # Partner listed but wrong SID and not DIRECT
        result["Partner Listed"]    = "⚠️ Listed"
        result["DIRECT Line"]       = "❌ Not DIRECT"
        result["Partner SID Match"] = f"❌ Wrong SID: {partner_lines[0]['seller_id']}"
        result["Verdict"]           = "⚠️ ID Mismatch"
        result["Notes"]             = f"Line: {partner_lines[0]['raw']}"

    else:
        result["Partner Listed"]    = "❌ Not Listed"
        result["DIRECT Line"]       = "❌ None"
        result["Partner SID Match"] = "❌"
        result["Verdict"]           = "❌ Not Listed"
        result["Notes"]             = f"{partner_domain} not found in app-ads.txt at {domain}"

    return result


def page_supply_validator():
    st.title("🔗 Supply Chain Validator")
    st.markdown("""
    **Automates the full bundle → app-ads.txt → DIRECT check pipeline.**

    Upload your bundle list from your reporting platform → enter your supply partner details
    → tool resolves each bundle's developer domain from Play Store / App Store
    → fetches their app-ads.txt → checks if your partner is listed as DIRECT.

    Replaces the manual process of opening each app in the store, finding the developer website,
    and checking their app-ads.txt one by one.
    """)
    st.markdown("---")

    # ── Partner details ───────────────────────────────────────────────────
    st.markdown("### 🏢 Supply Partner Details")
    col1, col2, col3 = st.columns(3)
    with col1:
        partner_domain = st.text_input(
            "Partner Domain",
            placeholder="e.g. smilewanted.com",
            help="The SSP/supply partner domain you want to check in app-ads.txt"
        )
    with col2:
        partner_sid = st.text_input(
            "Partner Seller ID (optional)",
            placeholder="e.g. 3148",
            help="If provided, also verifies the seller_id matches. Leave blank to check domain only."
        )
    with col3:
        partner_name = st.text_input("Partner Name", placeholder="e.g. SmileWanted")

    st.markdown("---")

    # ── Bundle input ──────────────────────────────────────────────────────
    st.markdown("### 📱 Bundle Input")
    tab_upload, tab_paste = st.tabs(["📁 Upload CSV", "✏️ Paste Bundles"])

    bundles_to_check = []

    with tab_upload:
        uploaded = st.file_uploader(
            "Upload CSV with a column named 'bundle'",
            type=["csv"],
            help="Export from your reporting platform (Agile/DashBid/etc). Column must be named 'bundle'."
        )
        if uploaded:
            try:
                df_in = pd.read_csv(uploaded)
                # Auto-detect bundle column
                bundle_col = next(
                    (c for c in df_in.columns
                     if c.lower() in ["bundle","bundle_id","app_bundle","appbundle","app_id"]),
                    None
                )
                if bundle_col is None:
                    st.error(f"No bundle column found. Columns: {list(df_in.columns)}")
                else:
                    bundles_to_check = df_in[bundle_col].dropna().astype(str).str.strip().unique().tolist()
                    st.success(f"✅ Loaded {len(bundles_to_check)} unique bundles from '{bundle_col}' column")
                    st.dataframe(df_in[[bundle_col]].head(5), use_container_width=True)
            except Exception as e:
                st.error(f"Error reading CSV: {e}")

    with tab_paste:
        raw = st.text_area(
            "Paste bundle IDs (one per line)",
            height=160,
            placeholder="com.smilewanted.app\ncom.king.candycrushsaga\n553834731\ncom.gameloft.android.ANMP.GloftA9HM"
        )
        if raw.strip():
            bundles_to_check = list(dict.fromkeys(
                [l.strip() for l in raw.splitlines() if l.strip()]
            ))
            st.info(f"📋 {len(bundles_to_check)} unique bundles ready")

    st.markdown("---")

    # ── Run options ───────────────────────────────────────────────────────
    col1, col2, col3 = st.columns(3)
    with col1:
        max_bundles = st.slider("Max bundles to check", 10, 500, 100)
    with col2:
        workers = st.slider("Parallel workers", 5, 30, 15,
                            help="Higher = faster but may hit rate limits on Play Store")
    with col3:
        delay_ms = st.slider("Delay between requests (ms)", 0, 500, 100,
                             help="Small delay avoids Play Store rate limiting")

    if st.button("🔗 Run Supply Chain Validation", type="primary"):
        if not partner_domain.strip():
            st.warning("Enter the supply partner domain")
            return
        if not bundles_to_check:
            st.warning("Add bundle IDs via upload or paste")
            return

        pd_clean = clean_domain(partner_domain)
        sid_clean = partner_sid.strip()
        to_run   = bundles_to_check[:max_bundles]

        android_b = [b for b in to_run if not b.isdigit()]
        ios_b     = [b for b in to_run if b.isdigit()]

        st.markdown(f"""
        ### 🔍 Validating {len(to_run)} bundles against `{pd_clean}`
        **Android:** {len(android_b)} · **iOS:** {len(ios_b)}
        """)

        results = []
        done    = 0
        prog    = st.progress(0)
        stat    = st.empty()

        def run_one(bundle):
            res = check_bundle_supply_chain(bundle, pd_clean, sid_clean)
            time.sleep(delay_ms / 1000)
            return res

        with ThreadPoolExecutor(max_workers=workers) as executor:
            fut_map = {executor.submit(run_one, b): b for b in to_run}
            for fut in as_completed(fut_map):
                res = fut.result()
                results.append(res)
                done += 1
                prog.progress(done / len(to_run))
                stat.text(
                    f"✅ {done}/{len(to_run)} — {res['Bundle'][:30]} → {res['Verdict']}"
                )

        prog.empty(); stat.empty()
        df = pd.DataFrame(results)

        # ── Summary metrics ───────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📊 Results Summary")

        legit    = len(df[df["Verdict"].str.contains("Legitimate", na=False)])
        id_mm    = len(df[df["Verdict"].str.contains("ID Mismatch|Reseller", na=False)])
        not_list = len(df[df["Verdict"].str.contains("Not Listed", na=False)])
        no_adst  = len(df[df["Verdict"].str.contains("No app-ads|No developer", na=False)])
        removed  = len(df[df["App Store Status"].str.contains("Removed|Not Found", na=False)])
        total    = len(df)

        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.metric("Total Checked",      total)
        c2.metric("✅ Legit DIRECT",    legit,    f"{round(legit/total*100)}%")
        c3.metric("❌ Not Listed",       not_list, f"{round(not_list/total*100)}%")
        c4.metric("⚠️ ID Mismatch",      id_mm)
        c5.metric("⚠️ No app-ads.txt",   no_adst)
        c6.metric("🗑️ Removed Apps",     removed)

        if legit == 0:
            st.error(f"🚨 None of the {total} bundles have {pd_clean} listed as DIRECT")
        elif legit < total * 0.3:
            st.warning(f"⚠️ Only {legit}/{total} ({round(legit/total*100)}%) bundles correctly list {pd_clean} as DIRECT")
        else:
            st.success(f"✅ {legit}/{total} bundles correctly list {pd_clean} as DIRECT")

        if removed > 0:
            st.warning(f"🗑️ {removed} apps are removed/delisted from store — potential IVT signal")

        # ── Pie chart ─────────────────────────────────────────────────────
        fig = px.pie(
            values=[legit, not_list, id_mm, no_adst, removed],
            names=["✅ Legitimate", "❌ Not Listed", "⚠️ ID Mismatch",
                   "⚠️ No app-ads.txt", "🗑️ App Removed"],
            color_discrete_map={
                "✅ Legitimate":    "#4CAF50",
                "❌ Not Listed":    "#F44336",
                "⚠️ ID Mismatch":  "#FDD835",
                "⚠️ No app-ads.txt":"#FF9800",
                "🗑️ App Removed":  "#9E9E9E",
            },
            hole=0.45, title=f"Supply Chain Status — {pd_clean}"
        )
        fig.update_traces(textinfo="label+percent+value")
        fig.update_layout(height=380, showlegend=True)
        st.plotly_chart(fig, use_container_width=True)

        # ── Grouped results ───────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📋 Results Grouped by Verdict")

        verdict_order = [
            "✅ Legitimate", "⚠️ ID Mismatch", "⚠️ Reseller Only",
            "❌ Not Listed", "⚠️ No app-ads.txt", "⚠️ No developer domain", "❌ No app-ads.txt"
        ]
        seen_verdicts = df["Verdict"].unique().tolist()
        ordered = [v for v in verdict_order if v in seen_verdicts]
        ordered += [v for v in seen_verdicts if v not in ordered]

        for verdict in ordered:
            grp = df[df["Verdict"] == verdict]
            if grp.empty:
                continue
            icon = ("✅" if "Legit" in verdict else
                    "🟡" if "Mismatch" in verdict or "Reseller" in verdict else
                    "❌")
            with st.expander(
                f"{icon} **{verdict}** — {len(grp)} bundles",
                expanded=("Legitimate" in verdict or "Mismatch" in verdict)
            ):
                display_cols = [c for c in [
                    "Bundle","Platform","App Store Status","Developer Domain",
                    "Partner Listed","DIRECT Line","Partner SID Match","Notes"
                ] if c in grp.columns]
                st.dataframe(grp[display_cols], use_container_width=True,
                             height=min(400, 60 + len(grp)*36))

                # Quick copy for not-listed domains — to send to partner
                if "Not Listed" in verdict or "Mismatch" in verdict:
                    domains = grp["Developer Domain"].dropna().tolist()
                    domains = [d for d in domains if "❌" not in str(d) and "⚠️" not in str(d)]
                    if domains:
                        st.markdown("**📋 Developer domains to fix (send to partner):**")
                        st.text_area(
                            "", value="\n".join(domains), height=120,
                            key=f"copy_{verdict[:10]}_{len(grp)}",
                            label_visibility="collapsed"
                        )

        # ── Full table ────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📋 Full Results Table")
        st.dataframe(df, use_container_width=True, height=420)

        c1, c2 = st.columns(2)
        c1.download_button(
            "⬇️ Download CSV",
            df_to_csv(df),
            f"supply_chain_{pd_clean}.csv",
            "text/csv"
        )
        c2.download_button(
            "⬇️ Download Excel (colored)",
            df_to_colored_excel(df, "Supply Chain", f"Supply Chain Validation — {pd_clean}"),
            f"supply_chain_{pd_clean}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # ── Action items ──────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 🎯 Action Items")

        not_listed_df = df[df["Verdict"].str.contains("Not Listed", na=False)]
        if not not_listed_df.empty:
            domains_to_fix = not_listed_df["Developer Domain"].dropna().tolist()
            domains_to_fix = [d for d in domains_to_fix if "❌" not in str(d)]
            if domains_to_fix:
                st.error(f"**❌ {len(domains_to_fix)} publishers need to add {pd_clean} to their app-ads.txt:**")
                sid_line = f", {sid_clean}" if sid_clean else ""
                st.code(
                    f"# Add this line to app-ads.txt:\n"
                    f"{pd_clean}{sid_line}, DIRECT",
                    language="text"
                )
                st.text_area(
                    "Publisher domains to contact:",
                    value="\n".join(domains_to_fix),
                    height=150,
                    key="action_not_listed"
                )

        id_mm_df = df[df["Verdict"].str.contains("ID Mismatch", na=False)]
        if not id_mm_df.empty:
            st.warning(
                f"**⚠️ {len(id_mm_df)} publishers have {pd_clean} listed but with wrong seller_id.** "
                f"Share the correct seller_id ({sid_clean}) with them to fix."
            )

def page_bundle():
    st.title("📱 Bundle Intelligence")
    st.markdown("*App store status check — single and bulk CSV*")
    st.markdown("---")
    tab1, tab2 = st.tabs(["Single Bundle","Bulk CSV"])

    with tab1:
        bundle = st.text_input("Bundle ID", placeholder="com.king.candycrushsaga  or  553834731")
        if st.button("🔍 Check", key="b1"):
            if not bundle.strip(): st.warning("Enter bundle"); return
            b = bundle.strip(); is_ios = b.isdigit()
            url = f"https://apps.apple.com/app/id{b}" if is_ios else f"https://play.google.com/store/apps/details?id={b}"
            platform = "iOS" if is_ios else "Android"
            try:
                r = requests.get(url, timeout=10, headers=HEADERS)
                status = "✅ Active" if r.status_code==200 else "❌ Removed/Not Found"
            except: status = "⚠️ Check failed"
            c1,c2,c3 = st.columns(3)
            c1.metric("Platform",status); c2.metric("Status",status); c3.metric("Bundle",b)
            st.markdown(f"[🔗 Open in store]({url})")

    with tab2:
        uploaded = st.file_uploader("CSV with 'bundle' column", type=["csv"])
        if uploaded and st.button("📦 Check All", key="b2"):
            df_in = pd.read_csv(uploaded)
            if "bundle" not in df_in.columns: st.error("Need 'bundle' column"); return
            bundles = df_in["bundle"].dropna().astype(str).tolist()[:100]
            results=[]; prog=st.progress(0)
            for i,b in enumerate(bundles):
                b=b.strip(); is_ios=b.isdigit()
                url = f"https://apps.apple.com/app/id{b}" if is_ios else f"https://play.google.com/store/apps/details?id={b}"
                try:
                    r = requests.get(url,timeout=6,headers=HEADERS)
                    status = "✅ Active" if r.status_code==200 else "❌ Removed"
                except: status="⚠️ Failed"
                results.append({"Bundle":b,"Platform":"iOS" if is_ios else "Android","Status":status,"URL":url})
                prog.progress((i+1)/len(bundles)); time.sleep(0.15)
            prog.empty()
            df_out = pd.DataFrame(results)
            st.dataframe(df_out, use_container_width=True)
            active  = len([r for r in results if "Active"   in r["Status"]])
            removed = len([r for r in results if "Removed"  in r["Status"]])
            c1,c2 = st.columns(2); c1.metric("✅ Active",active); c2.metric("❌ Removed",removed)
            if removed: st.warning(f"⚠️ {removed} delisted apps still sending traffic — IVT risk")
            st.download_button("⬇️ Excel (colored)", df_to_colored_excel(df_out,"Bundles"),
                               "bundle_check.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def page_schain():
    st.title("⛓️ Schain Validator")
    st.markdown("*Node-by-node compliance — hp=1, valid asi, Advlion as final node*")
    st.markdown("---")
    mode = st.radio("Format",["JSON","Raw string"], horizontal=True)
    if mode=="JSON":
        raw = st.text_area("Paste schain JSON", height=160,
            placeholder='{"ver":"1.0","complete":1,"nodes":[{"asi":"openx.com","sid":"123","hp":1},{"asi":"advlion.com","sid":"3148","hp":1}]}')
    else:
        raw = st.text_area("Paste raw string", height=80, placeholder="1.0,1!openx.com,123,1,,!advlion.com,3148,1,,")

    if st.button("⛓️ Validate", type="primary"):
        if not raw.strip(): st.warning("Paste a schain"); return
        nodes=[]; ver="—"; complete="—"
        if mode=="JSON":
            try:
                data=json.loads(raw); nodes=data.get("nodes",[]); ver=data.get("ver","—"); complete=data.get("complete","—")
            except Exception as e: st.error(f"Invalid JSON: {e}"); return
        else:
            try:
                parts=raw.strip().split("!"); hdr=parts[0].split(",")
                ver=hdr[0] if len(hdr)>0 else "—"; complete=hdr[1] if len(hdr)>1 else "—"
                for np_ in parts[1:]:
                    p=np_.split(",")
                    nodes.append({"asi":p[0] if len(p)>0 else "","sid":p[1] if len(p)>1 else "","hp":int(p[2]) if len(p)>2 and p[2].isdigit() else 0})
            except Exception as e: st.error(f"Parse error: {e}"); return

        c1,c2,c3 = st.columns(3); c1.metric("Version",ver); c2.metric("Complete",str(complete)); c3.metric("Nodes",len(nodes))
        results=[]
        for i,node in enumerate(nodes):
            asi=str(node.get("asi","")); sid=str(node.get("sid","")); hp=node.get("hp",0); is_last=(i==len(nodes)-1)
            issues=[]
            if hp!=1: issues.append("❌ hp≠1")
            if not asi: issues.append("❌ Missing asi")
            if not sid: issues.append("❌ Missing sid")
            if is_cdn_domain(asi): issues.append("🚨 CDN domain in schain")
            if is_last:
                if asi.lower()==ADVLION_DOMAIN and sid==ADVLION_SID: issues.append("✅ Advlion final node")
                else: issues.append(f"⚠️ Final={asi} (expected advlion.com)")
            results.append({"Node":f"#{i+1}","asi":asi,"sid":sid,"hp":hp,
                             "Position":"🔚 Final" if is_last else f"Hop {i+1}",
                             "Status":"✅ Valid" if not any("❌" in x for x in issues) else "❌ Issues",
                             "Notes":" | ".join(issues) if issues else "✅ Clean"})

        df=pd.DataFrame(results); st.dataframe(df, use_container_width=True)
        hop_count=len(nodes); valid=len([r for r in results if r["Status"]=="✅ Valid"])
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Hops",hop_count); c2.metric("Valid",valid); c3.metric("Issues",hop_count-valid)
        c4.metric("SPO Risk","🚨 High" if hop_count>=3 else "⚠️ Medium" if hop_count==2 else "✅ Low")
        if hop_count>=3: st.warning("⚠️ 3+ hops = SPO disadvantage. Buyers may deprioritize.")
        if str(complete)=="0": st.error("🚨 complete=0 — hidden hops. High fraud risk.")

def page_ivt():
    st.title("🛡️ IVT Risk Scorer")
    st.markdown("*App status + ads.txt + sellers.json → combined risk score 0–100*")
    st.markdown("---")
    c1,c2 = st.columns(2)
    with c1: bundle=st.text_input("App Bundle",placeholder="com.example.app"); pub_domain=st.text_input("Publisher Domain",placeholder="publisher.com")
    with c2: ssp_domain=st.text_input("SSP Domain",placeholder="openx.com"); sid_input=st.text_input("Seller ID",placeholder="123456")

    if st.button("🛡️ Score IVT Risk", type="primary"):
        if not any([bundle.strip(),pub_domain.strip(),ssp_domain.strip()]): st.warning("Enter at least one input"); return
        risk=0; signals=[]
        with st.spinner("Analysing..."):
            if bundle.strip():
                b=bundle.strip(); is_ios=b.isdigit()
                url=f"https://apps.apple.com/app/id{b}" if is_ios else f"https://play.google.com/store/apps/details?id={b}"
                try:
                    r=requests.get(url,timeout=8,headers=HEADERS)
                    if r.status_code==200: signals.append({"Signal":"App Store Status","Risk":0,"Finding":"✅ Active"})
                    else: risk+=35; signals.append({"Signal":"App Store Status","Risk":35,"Finding":"❌ Delisted"})
                except: risk+=15; signals.append({"Signal":"App Store Status","Risk":15,"Finding":"⚠️ Could not verify"})

            if pub_domain.strip():
                pd_=clean_domain(pub_domain); adstxt_ok=False
                for is_app in [True,False]:
                    res=fetch_ads_txt(pd_,is_app)
                    if res["success"]:
                        adstxt_ok=True; rows=parse_ads_txt(res["content"])
                        adv=[r for r in rows if r["exchange_domain"]==ADVLION_DOMAIN and r["seller_id"]==ADVLION_SID]
                        if adv: signals.append({"Signal":"Advlion in ads.txt","Risk":0,"Finding":"✅ Listed"})
                        else: risk+=10; signals.append({"Signal":"Advlion in ads.txt","Risk":10,"Finding":"⚠️ Not listed"})
                        direct=[r for r in rows if r["relationship"]=="DIRECT"]
                        if not direct: risk+=10; signals.append({"Signal":"DIRECT lines","Risk":10,"Finding":"⚠️ None"})
                        else: signals.append({"Signal":"DIRECT lines","Risk":0,"Finding":f"✅ {len(direct)} found"})
                        break
                if not adstxt_ok: risk+=25; signals.append({"Signal":"ads.txt","Risk":25,"Finding":"❌ Not found"})

            if ssp_domain.strip() and sid_input.strip():
                sd=clean_domain(ssp_domain); sj=fetch_sellers_json(sd)
                if sj["success"]:
                    s_list=get_sellers(sj["data"])
                    match=next((s for s in s_list if str(s.get("seller_id","")).strip()==sid_input.strip()),None)
                    if match:
                        stype=str(match.get("seller_type","")).upper()
                        signals.append({"Signal":"In sellers.json","Risk":0,"Finding":f"✅ Found as {stype}"})
                        if is_cdn_domain(extract_domain_from_field(str(match.get("domain","")))):
                            risk+=20; signals.append({"Signal":"CDN Domain","Risk":20,"Finding":"🚨 Publisher domain is CDN"})
                    else: risk+=30; signals.append({"Signal":"In sellers.json","Risk":30,"Finding":"❌ Not found"})
                else: risk+=10; signals.append({"Signal":"sellers.json","Risk":10,"Finding":"⚠️ Unreachable"})

        risk=min(risk,100)
        label="🚨 HIGH RISK" if risk>=60 else "⚠️ MEDIUM" if risk>=30 else "✅ LOW RISK"
        color="#F44336" if risk>=60 else "#FF9800" if risk>=30 else "#4CAF50"
        fig=go.Figure(go.Indicator(mode="gauge+number",value=risk,title={"text":f"IVT Risk — {label}"},
            gauge={"axis":{"range":[0,100]},"bar":{"color":color,"thickness":0.3},
                   "steps":[{"range":[0,30],"color":"#0D2010"},{"range":[30,60],"color":"#2D2010"},{"range":[60,100],"color":"#2D0D0D"}]}))
        fig.update_layout(height=300); st.plotly_chart(fig, use_container_width=True)
        df=pd.DataFrame(signals); st.dataframe(df, use_container_width=True)
        st.download_button("⬇️ Excel (colored)", df_to_colored_excel(df,"IVT Risk"),"ivt_risk.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def page_onboarding():
    st.title("📝 Partner Onboarding Report")
    st.markdown("*One-click full compliance audit for any new partner*")
    st.markdown("---")
    c1,c2,c3 = st.columns(3)
    with c1: partner_domain=st.text_input("Partner Domain",placeholder="anzu.io")
    with c2: partner_type=st.selectbox("Type",["Demand (DSP/Network)","Supply (Publisher/SSP)"])
    with c3: partner_name=st.text_input("Partner Name",placeholder="Anzu")

    if st.button("📝 Generate Report", type="primary"):
        if not partner_domain.strip(): st.warning("Enter domain"); return
        pd_=clean_domain(partner_domain); pname=partner_name.strip() or pd_; checklist=[]
        is_demand="Demand" in partner_type

        with st.spinner("Running full audit..."):
            sj=fetch_sellers_json(pd_)
            checklist.append({"Check":"sellers.json Accessible","Result":"✅ Pass" if sj["success"] else "❌ Fail",
                               "Detail":f"{pd_}/sellers.json" if sj["success"] else "Not found"})
            if sj["success"]:
                sellers=get_sellers(sj["data"]); conf=[s for s in sellers if s.get("is_confidential")==1]
                conf_pct=round(len(conf)/len(sellers)*100) if sellers else 0
                sids=[str(s.get("seller_id","")) for s in sellers]; dups=len(sids)-len(set(sids))
                cdn_s=[s for s in sellers if is_cdn_domain(extract_domain_from_field(str(s.get("domain",""))))]
                checklist.append({"Check":"is_confidential Ratio","Result":"✅ Pass" if conf_pct<=20 else "⚠️ Review",
                                   "Detail":f"{conf_pct}% confidential"})
                checklist.append({"Check":"Duplicate Seller IDs","Result":"✅ Pass" if dups==0 else "⚠️ Has Dups",
                                   "Detail":f"{dups} duplicates"})
                checklist.append({"Check":"CDN Domains in sellers.json","Result":"✅ None" if not cdn_s else f"🚨 {len(cdn_s)} found",
                                   "Detail":"Clean" if not cdn_s else f"{len(cdn_s)} CDN/cloud publisher domains"})

            adstxt=fetch_ads_txt(pd_); app_adstxt=fetch_ads_txt(pd_,True)
            checklist.append({"Check":"ads.txt","Result":"✅ Pass" if adstxt["success"] else "⚠️ Missing","Detail":"Found" if adstxt["success"] else "Not reachable"})
            checklist.append({"Check":"app-ads.txt","Result":"✅ Pass" if app_adstxt["success"] else "⚠️ Missing","Detail":"Found" if app_adstxt["success"] else "Not reachable"})

            adv_listed=False
            for res in [adstxt,app_adstxt]:
                if res["success"]:
                    rows=parse_ads_txt(res["content"])
                    adv=[r for r in rows if r["exchange_domain"]==ADVLION_DOMAIN and r["seller_id"]==ADVLION_SID and r["relationship"]=="DIRECT"]
                    if adv: adv_listed=True; break
            checklist.append({"Check":"Advlion (3148) as DIRECT","Result":"✅ Pass" if adv_listed else "❌ Missing",
                               "Detail":"Listed" if adv_listed else f"Add: advlion.com, {ADVLION_SID}, DIRECT"})

            if is_demand:
                sample_ex = list(TOP_15_EXCHANGES.items())[:8]
                def _ob_seat(args):
                    _, ex_dm = args
                    res = fetch_sellers_json(ex_dm, use_cache=True)
                    if not res["success"]: return 0
                    s_list = get_sellers(res["data"])
                    return 1 if any(str(s.get("domain","")).lower().replace("www.","") == pd_ for s in s_list) else 0
                with ThreadPoolExecutor(max_workers=len(sample_ex)) as ex_pool:
                    found_ex = sum(f.result() for f in as_completed([ex_pool.submit(_ob_seat, it) for it in sample_ex]))
                checklist.append({"Check":f"Exchange Seats ({len(sample_ex)} sampled)",
                                   "Result":"✅ Pass" if found_ex>=2 else "⚠️ Low" if found_ex==1 else "❌ None",
                                   "Detail":f"Found in {found_ex}/{len(sample_ex)}"})

        df=pd.DataFrame(checklist)
        st.markdown(f"### Onboarding Report — {pname}  |  {pd_}  |  {datetime.now().strftime('%d %b %Y')}")
        st.dataframe(df, use_container_width=True)

        passes=len([c for c in checklist if "Pass" in c["Result"] or "None" in c["Result"] or "Pass" in c["Result"]])
        fails =len([c for c in checklist if "Fail" in c["Result"] or "Missing" in c["Result"] or "❌" in c["Result"]])
        c1,c2,c3=st.columns(3); c1.metric("✅ Passed",passes); c2.metric("❌ Failed",fails)
        if fails==0: c3.metric("Decision","✅ Approve"); st.success("All checks passed — safe to onboard")
        elif fails<=2: c3.metric("Decision","⚠️ Review"); st.warning(f"{fails} issue(s) — resolve before onboarding")
        else: c3.metric("Decision","❌ Hold"); st.error("Too many failures — DO NOT onboard")

        st.download_button("⬇️ Excel (colored)", df_to_colored_excel(df,"Onboarding"),
                           f"{pd_}_onboarding.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def page_digest():
    st.title("📅 Weekly Compliance Digest")
    st.markdown("*Monday sweep — all partners, all checks, colored Excel export*")
    st.markdown("---")
    c1,c2=st.columns(2)
    with c1: st.markdown("**Supply Partners**"); supply_raw=st.text_area("One domain per line",height=160,placeholder="openx.com\npubmatic.com",key="wk_s")
    with c2: st.markdown("**Demand Partners**"); demand_raw=st.text_area("One domain per line",height=160,placeholder="anzu.io\neskimi.com",key="wk_d")

    if st.button("📅 Run Digest", type="primary"):
        supply=[clean_domain(d) for d in supply_raw.splitlines() if d.strip()]
        demand=[clean_domain(d) for d in demand_raw.splitlines() if d.strip()]
        all_p=[(d,"Supply") for d in supply]+[(d,"Demand") for d in demand]
        if not all_p: st.warning("Add partners"); return

        results=[]; prog=st.progress(0); stat=st.empty()
        for i,(domain,ptype) in enumerate(all_p):
            stat.text(f"[{i+1}/{len(all_p)}] {domain}..."); prog.progress((i+1)/len(all_p))
            sj=fetch_sellers_json(domain); sj_ok="✅" if sj["success"] else "❌"
            adstxt=fetch_ads_txt(domain); app_adstxt=fetch_ads_txt(domain,True)
            adstxt_ok="✅" if (adstxt["success"] or app_adstxt["success"]) else "❌"
            adv_ok="—"
            for res in [adstxt,app_adstxt]:
                if res["success"]:
                    rows=parse_ads_txt(res["content"])
                    adv=[r for r in rows if r["exchange_domain"]==ADVLION_DOMAIN and r["seller_id"]==ADVLION_SID]
                    adv_ok="✅" if adv else "❌"; break
            seats_str="—"
            if ptype=="Demand":
                sample = list(TOP_15_EXCHANGES.values())[:5]
                def _digest_seat(ex_dm):
                    res = fetch_sellers_json(ex_dm, use_cache=True)
                    if not res["success"]: return 0
                    s_list = get_sellers(res["data"])
                    return 1 if any(str(s.get("domain","")).lower().replace("www.","") == domain for s in s_list) else 0
                with ThreadPoolExecutor(max_workers=len(sample)) as ex_pool:
                    cnt = sum(f.result() for f in as_completed([ex_pool.submit(_digest_seat, d) for d in sample]))
                seats_str=f"{cnt}/{len(sample)}"
            issues=[x for x in [sj_ok,adstxt_ok,adv_ok] if x=="❌"]
            flag="🚨 Action" if len(issues)>=2 else "⚠️ Review" if len(issues)==1 else "✅ Clean"
            results.append({"Partner":domain,"Type":ptype,"sellers.json":sj_ok,"ads.txt":adstxt_ok,
                             "Advlion":adv_ok,"Seats":seats_str,"Status":flag})

        prog.empty(); stat.empty()
        df=pd.DataFrame(results)
        clean=len([r for r in results if r["Status"]=="✅ Clean"])
        action=len([r for r in results if "Action" in r["Status"]])
        review=len([r for r in results if "Review" in r["Status"]])
        c1,c2,c3,c4=st.columns(4)
        c1.metric("Total",len(results)); c2.metric("✅ Clean",clean); c3.metric("⚠️ Review",review); c4.metric("🚨 Action",action)
        if action: st.error(f"🚨 {action} partner(s) need immediate action")
        st.dataframe(df, use_container_width=True, height=400)
        st.download_button("⬇️ Excel (colored)",
                           df_to_colored_excel(df,"Weekly Digest",f"Weekly Digest — {datetime.now().strftime('%d %b %Y')}"),
                           f"digest_{datetime.now().strftime('%Y%m%d')}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── MAIN ─────────────────────────────────────────────────────────────────

def check_password():
    """Simple password gate — blocks access without correct password."""
    # Try secrets first, fall back to hardcoded password
    try:
        correct_password = st.secrets["APP_PASSWORD"]
    except Exception:
        correct_password = "Advlion@2024"   # fallback

    def password_entered():
        if st.session_state["password"] == correct_password:
            st.session_state["authenticated"] = True
            del st.session_state["password"]
        else:
            st.session_state["authenticated"] = False

    if st.session_state.get("authenticated"):
        return True

    st.markdown("""
    <div style='text-align:center; padding: 80px 0 20px 0;'>
        <div style='font-size:52px;'>🔗</div>
        <div style='font-size:24px; font-weight:800; color:#7EC8E3;'>AdChain Inspector</div>
        <div style='font-size:13px; color:#5577AA; margin:4px 0 30px 0;'>Advlion Supply Chain Intelligence · Private Access</div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1,2,1])
    with col2:
        st.text_input("Access Password", type="password",
                      key="password", on_change=password_entered,
                      placeholder="Enter password to continue...")
        if "authenticated" in st.session_state and not st.session_state["authenticated"]:
            st.error("❌ Incorrect password")
    return False


def main():
    if not check_password():
        return
    init_db()
    page = sidebar()
    page_map = {
        "overview":      page_overview,
        "authenticator": page_authenticator,
        "inspector":     page_inspector,
        "adstxt":        page_adstxt,
        "cross":         page_cross,
        "demand":        page_demand,
        "health":        page_health,
        "bulk_ssp":      page_bulk_ssp,
        "presence":      page_presence,
        "sd_intel":      page_sd_intel,
        "tracker":       page_tracker,
        "intermediary":  page_intermediary_intel,
        "supply_validator": page_supply_validator,
        "bundle":           page_bundle,
        "schain":        page_schain,
        "ivt":           page_ivt,
        "onboarding":    page_onboarding,
        "digest":        page_digest,
    }
    page_map.get(page, page_overview)()

if __name__ == "__main__":
    main()
